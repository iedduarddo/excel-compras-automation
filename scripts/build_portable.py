"""Constrói e valida a distribuição portátil do projeto para Windows x64."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import platform
import re
import shutil
import subprocess
import sys
import tempfile
import tomllib
from datetime import date
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

from openpyxl import Workbook, load_workbook

APP_NAME = "ExcelComprasAutomation"
PLATFORM_NAME = "windows-x64"
ZIP_TIMESTAMP = (2020, 1, 1, 0, 0, 0)
OPERATING_DIRECTORIES = ("input", "output", "backup", "logs")
REQUIRED_TOP_LEVEL = {
    "_internal",
    "backup",
    "CHANGELOG.md",
    "config",
    "docs",
    f"{APP_NAME}.exe",
    "iniciar.cmd",
    "input",
    "LEIA-ME.txt",
    "logs",
    "output",
    "README_PORTATIL.md",
    "run.ps1",
    "VERSAO.txt",
}
REQUIRED_DOCS = {
    "ADAPTADORES.md",
    "ASSISTENTE.md",
    "SOLUCAO_DE_PROBLEMAS.md",
}
FORBIDDEN_DIRECTORY_NAMES = {
    ".git",
    ".pytest_cache",
    ".venv",
    "__pycache__",
    "tests",
}
FORBIDDEN_FILE_SUFFIXES = {
    ".env",
    ".key",
    ".pem",
    ".xlsm",
    ".xlsx",
}
VERSION_PATTERN = re.compile(r"^\d+\.\d+\.\d+$")
COMMIT_PATTERN = re.compile(r"^[0-9a-fA-F]{40}(?:[0-9a-fA-F]{24})?$")


def read_project_version(project_root: Path) -> str:
    """Confere e devolve a versão sincronizada do projeto."""

    pyproject_path = project_root / "pyproject.toml"
    package_path = project_root / "src" / "__init__.py"
    project_data = tomllib.loads(pyproject_path.read_text(encoding="utf-8"))
    project_version = project_data["project"]["version"]
    package_source = package_path.read_text(encoding="utf-8")
    package_match = re.search(
        r'^__version__\s*=\s*["\'](?P<version>[^"\']+)["\']\s*$',
        package_source,
        flags=re.MULTILINE,
    )

    if not isinstance(project_version, str) or package_match is None:
        raise ValueError("Não foi possível ler as duas versões do projeto.")

    package_version = package_match.group("version")
    if project_version != package_version:
        raise ValueError(
            "As versões do projeto não estão sincronizadas: "
            f"pyproject={project_version}; pacote={package_version}."
        )
    if VERSION_PATTERN.fullmatch(project_version) is None:
        raise ValueError(
            "A distribuição portátil exige uma versão estável no formato X.Y.Z."
        )
    return project_version


def portable_package_name(version: str) -> str:
    """Produz o nome estável do diretório e do ZIP."""

    if VERSION_PATTERN.fullmatch(version) is None:
        raise ValueError(f"Versão inválida para o pacote portátil: {version!r}.")
    return f"{APP_NAME}-v{version}-{PLATFORM_NAME}"


def version_quad(version: str) -> tuple[int, int, int, int]:
    """Converte uma versão X.Y.Z no formato exigido pelo Windows."""

    if VERSION_PATTERN.fullmatch(version) is None:
        raise ValueError(f"Versão inválida: {version!r}.")
    major, minor, patch = (int(part) for part in version.split("."))
    return major, minor, patch, 0


def normalize_commit(commit: str | None) -> str | None:
    """Valida o identificador Git registrado dentro do pacote."""

    if commit is None:
        return None
    normalized = commit.strip().casefold()
    if COMMIT_PATTERN.fullmatch(normalized) is None:
        raise ValueError("O commit deve ser um SHA Git hexadecimal completo.")
    return normalized


def render_version_resource(version: str) -> str:
    """Gera os metadados de versão consumidos pelo PyInstaller."""

    quad = version_quad(version)
    dotted_quad = ".".join(str(part) for part in quad)
    return f"""VSVersionInfo(
  ffi=FixedFileInfo(
    filevers={quad},
    prodvers={quad},
    mask=0x3f,
    flags=0x0,
    OS=0x40004,
    fileType=0x1,
    subtype=0x0,
    date=(0, 0)
  ),
  kids=[
    StringFileInfo([
      StringTable(
        '041604B0',
        [
          StringStruct('CompanyName', 'iedduarddo'),
          StringStruct('FileDescription', 'Excel Compras Automation'),
          StringStruct('FileVersion', '{dotted_quad}'),
          StringStruct('InternalName', '{APP_NAME}'),
          StringStruct('OriginalFilename', '{APP_NAME}.exe'),
          StringStruct('ProductName', 'Excel Compras Automation'),
          StringStruct('ProductVersion', '{version}')
        ]
      )
    ]),
    VarFileInfo([VarStruct('Translation', [1046, 1200])])
  ]
)
"""


def _safe_remove_directory(path: Path, *, expected_parent: Path) -> None:
    """Remove somente um filho direto previamente delimitado."""

    resolved = path.resolve()
    parent = expected_parent.resolve()
    if resolved == parent or resolved.parent != parent:
        raise ValueError(f"Diretório inseguro para limpeza: {resolved}")
    if resolved.exists():
        shutil.rmtree(resolved)


def _safe_remove_file(path: Path, *, expected_parent: Path) -> None:
    resolved = path.resolve()
    parent = expected_parent.resolve()
    if resolved.parent != parent:
        raise ValueError(f"Arquivo inseguro para limpeza: {resolved}")
    if resolved.exists():
        resolved.unlink()


def _run_command(
    command: list[str],
    *,
    cwd: Path,
    environment: dict[str, str] | None = None,
    timeout: int | None = None,
    capture_output: bool = False,
) -> subprocess.CompletedProcess[str]:
    completed = subprocess.run(  # noqa: S603
        command,
        cwd=cwd,
        env=environment,
        check=False,
        capture_output=capture_output,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=timeout,
    )
    if completed.returncode != 0:
        details = "\n".join(
            part.strip()
            for part in (completed.stdout, completed.stderr)
            if part and part.strip()
        )
        suffix = f"\n{details}" if details else ""
        raise RuntimeError(
            f"O comando terminou com código {completed.returncode}: "
            f"{' '.join(command)}{suffix}"
        )
    return completed


def _windows_batch_command(
    command_prompt: Path,
    batch_file: Path,
    *arguments: str,
) -> list[str]:
    """Monta uma chamada de batch segura para caminhos e argumentos com espaços."""

    return [
        str(command_prompt),
        "/d",
        "/s",
        "/v:off",
        "/c",
        "call",
        str(batch_file),
        *arguments,
    ]


def build_frozen_application(
    project_root: Path,
    scratch_directory: Path,
    version: str,
) -> Path:
    """Executa o PyInstaller por uma especificação versionada."""

    version_file = scratch_directory / "version_info.txt"
    version_file.parent.mkdir(parents=True, exist_ok=True)
    version_file.write_text(
        render_version_resource(version),
        encoding="utf-8",
        newline="\n",
    )

    distribution_directory = scratch_directory / "pyinstaller-dist"
    work_directory = scratch_directory / "pyinstaller-work"
    spec_file = project_root / "packaging" / f"{APP_NAME}.spec"
    if not spec_file.is_file():
        raise FileNotFoundError(f"Especificação do PyInstaller ausente: {spec_file}")

    _run_command(
        [
            sys.executable,
            "-m",
            "PyInstaller",
            "--noconfirm",
            "--clean",
            "--distpath",
            str(distribution_directory),
            "--workpath",
            str(work_directory),
            str(spec_file),
        ],
        cwd=project_root,
    )

    frozen_directory = distribution_directory / APP_NAME
    if not (frozen_directory / f"{APP_NAME}.exe").is_file():
        raise RuntimeError("O PyInstaller não produziu o executável esperado.")
    return frozen_directory


def assemble_portable_package(
    *,
    project_root: Path,
    frozen_directory: Path,
    package_directory: Path,
    version: str,
    commit: str | None,
) -> None:
    """Monta o pacote final por uma lista explícita de arquivos."""

    shutil.copytree(frozen_directory, package_directory)

    config_directory = package_directory / "config"
    config_directory.mkdir()
    for filename in ("aliases.json", "rules.json"):
        shutil.copy2(
            project_root / "config" / filename,
            config_directory / filename,
        )

    for directory_name in OPERATING_DIRECTORIES:
        (package_directory / directory_name).mkdir()

    for filename in ("iniciar.cmd", "run.ps1", "CHANGELOG.md"):
        shutil.copy2(project_root / filename, package_directory / filename)

    docs_directory = package_directory / "docs"
    docs_directory.mkdir()
    for filename in sorted(REQUIRED_DOCS):
        shutil.copy2(
            project_root / "docs" / filename,
            docs_directory / filename,
        )

    version_marker = f"v{version}"
    for source_name, destination_name in (
        ("README_PORTATIL.md", "README_PORTATIL.md"),
        ("LEIA-ME.txt", "LEIA-ME.txt"),
    ):
        guide = (project_root / "packaging" / source_name).read_text(encoding="utf-8")
        (package_directory / destination_name).write_text(
            guide.replace("vX.Y.Z", version_marker),
            encoding="utf-8",
            newline="\n",
        )

    commit_value = normalize_commit(commit) or "não informado"
    (package_directory / "VERSAO.txt").write_text(
        (
            f"Excel Compras Automation {version}\n"
            f"Plataforma: Windows x64\n"
            f"Commit: {commit_value}\n"
        ),
        encoding="utf-8",
        newline="\n",
    )


def validate_portable_package(package_directory: Path) -> None:
    """Rejeita conteúdo ausente, operacional ou fora do escopo."""

    actual_top_level = {path.name for path in package_directory.iterdir()}
    if actual_top_level != REQUIRED_TOP_LEVEL:
        missing = sorted(REQUIRED_TOP_LEVEL - actual_top_level)
        unexpected = sorted(actual_top_level - REQUIRED_TOP_LEVEL)
        raise ValueError(
            "Conteúdo de primeiro nível inválido. "
            f"Ausentes={missing}; inesperados={unexpected}."
        )

    config_files = {path.name for path in (package_directory / "config").iterdir()}
    if config_files != {"aliases.json", "rules.json"}:
        raise ValueError(f"Configuração externa inesperada: {sorted(config_files)}")
    for filename in config_files:
        json.loads(
            (package_directory / "config" / filename).read_text(encoding="utf-8")
        )

    docs_files = {path.name for path in (package_directory / "docs").iterdir()}
    if docs_files != REQUIRED_DOCS:
        raise ValueError(f"Documentação inesperada: {sorted(docs_files)}")

    for directory_name in OPERATING_DIRECTORIES:
        directory = package_directory / directory_name
        if not directory.is_dir() or any(directory.iterdir()):
            raise ValueError(
                f"A pasta operacional deve existir vazia: {directory_name}"
            )

    pythoncom_dlls: list[Path] = []
    pywintypes_dlls: list[Path] = []
    for path in package_directory.rglob("*"):
        relative_parts = {
            part.casefold() for part in path.relative_to(package_directory).parts
        }
        if relative_parts.intersection(FORBIDDEN_DIRECTORY_NAMES):
            raise ValueError(f"Diretório proibido no pacote: {path}")
        if not path.is_file():
            continue

        lower_name = path.name.casefold()
        if (
            lower_name in FORBIDDEN_FILE_SUFFIXES
            or path.suffix.casefold() in FORBIDDEN_FILE_SUFFIXES
            or lower_name.startswith(".env")
        ):
            raise ValueError(f"Arquivo proibido no pacote: {path}")
        if lower_name.startswith("pythoncom") and lower_name.endswith(".dll"):
            pythoncom_dlls.append(path)
        if lower_name.startswith("pywintypes") and lower_name.endswith(".dll"):
            pywintypes_dlls.append(path)

    if not pythoncom_dlls or not pywintypes_dlls:
        raise ValueError(
            "As DLLs obrigatórias do pywin32 não foram encontradas no pacote."
        )


def create_smoke_workbook(output_file: Path, aliases_file: Path) -> None:
    """Cria uma entrada sintética, sem dados reais, para o smoke test."""

    aliases = json.loads(aliases_file.read_text(encoding="utf-8"))
    workbook = Workbook()
    base = workbook.active
    base.title = aliases["sheets"]["base"][0]
    base_headers = [
        "request_id",
        "request_date",
        "travel_date",
        "traveler",
        "area",
        "cost_center",
        "service_type",
        "destination",
        "supplier",
        "quantity",
        "unit_value",
        "fees",
        "booking_status",
        "criticality",
        "card_status",
    ]
    base.append([aliases["base_columns"][field][0] for field in base_headers])
    base.append(
        [
            "SMOKE-001",
            date(2026, 1, 2),
            date(2026, 1, 15),
            "Pessoa de Teste",
            "Compras",
            "CC-SMOKE",
            "Aéreo",
            "Cidade de Teste",
            "Fornecedor Sintético",
            1,
            800,
            50,
            "Confirmada",
            "Normal",
            "Conferido",
        ]
    )

    policies = workbook.create_sheet(aliases["sheets"]["policies"][0])
    policy_headers = ("service_type", "limit_value", "min_lead_days")
    policies.append([aliases["policy_columns"][field][0] for field in policy_headers])
    policies.append(["Aéreo", 1000, 5])

    responses = workbook.create_sheet(aliases["sheets"]["responses"][0])
    response_headers = ("indicator", "answer", "formula_used")
    responses.append(
        [aliases["response_columns"][field][0] for field in response_headers]
    )
    for row, label_options in enumerate(
        aliases["indicator_labels"].values(),
        start=2,
    ):
        responses.cell(row, 1, label_options[0])

    for column, title in enumerate(
        [
            "ID Solicitação",
            "Motivo da prioridade",
            "Ação recomendada",
            "Ordem (1 a 5)",
        ],
        start=1,
    ):
        responses.cell(14, column, title)
    responses.cell(25, 1, "Tabela dinâmica e gráfico")

    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)
    workbook.close()


def _isolated_windows_environment() -> dict[str, str]:
    environment = os.environ.copy()
    for key in list(environment):
        if key.upper().startswith("PYTHON"):
            environment.pop(key)

    windows_directory = Path(environment.get("SystemRoot", r"C:\Windows"))
    environment["PATH"] = os.pathsep.join(
        [
            str(windows_directory / "System32"),
            str(windows_directory / "System32" / "WindowsPowerShell" / "v1.0"),
            str(windows_directory),
        ]
    )
    return environment


def run_portable_smoke_test(package_directory: Path, version: str) -> None:
    """Valida o executável fora do repositório e sem Python externo."""

    with tempfile.TemporaryDirectory(prefix="Excel Compras Portatil ") as temporary:
        smoke_parent = Path(temporary)
        smoke_package = smoke_parent / f"{package_directory.name} com espaços"
        shutil.copytree(package_directory, smoke_package)
        input_file = smoke_package / "input" / "entrada_sintetica.xlsx"
        create_smoke_workbook(
            input_file,
            smoke_package / "config" / "aliases.json",
        )
        original_hash = hashlib.sha256(input_file.read_bytes()).hexdigest()
        executable = (smoke_package / f"{APP_NAME}.exe").resolve()
        launcher = (smoke_package / "iniciar.cmd").resolve()
        run_script = (smoke_package / "run.ps1").resolve()
        environment = _isolated_windows_environment()
        external_cwd = smoke_parent / "diretorio de trabalho externo"
        external_cwd.mkdir()
        windows_directory = Path(environment.get("SystemRoot", r"C:\Windows"))
        command_prompt = windows_directory / "System32" / "cmd.exe"
        powershell = (
            windows_directory
            / "System32"
            / "WindowsPowerShell"
            / "v1.0"
            / "powershell.exe"
        )

        version_result = _run_command(
            [str(executable), "--version"],
            cwd=external_cwd,
            environment=environment,
            timeout=60,
            capture_output=True,
        )
        if version not in version_result.stdout:
            raise RuntimeError("O executável não informou a versão esperada.")

        launcher_version = _run_command(
            _windows_batch_command(
                command_prompt,
                launcher,
                "-Version",
            ),
            cwd=external_cwd,
            environment=environment,
            timeout=60,
            capture_output=True,
        )
        if version not in launcher_version.stdout:
            raise RuntimeError("O iniciador não encaminhou a versão esperada.")

        help_result = _run_command(
            [str(executable), "--help"],
            cwd=external_cwd,
            environment=environment,
            timeout=60,
            capture_output=True,
        )
        if "--diagnostico" not in help_result.stdout:
            raise RuntimeError("A ajuda do executável está incompleta.")

        diagnostic_result = _run_command(
            [
                str(powershell),
                "-NoLogo",
                "-NoProfile",
                "-NonInteractive",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(run_script),
                "-Diagnostico",
            ],
            cwd=external_cwd,
            environment=environment,
            timeout=90,
            capture_output=True,
        )
        if "AMBIENTE PRONTO" not in diagnostic_result.stdout:
            raise RuntimeError("O diagnóstico do pacote não aprovou o ambiente.")
        for directory_name in ("output", "backup", "logs"):
            if any((smoke_package / directory_name).iterdir()):
                raise RuntimeError("O diagnóstico criou um artefato operacional.")

        _run_command(
            _windows_batch_command(
                command_prompt,
                launcher,
                "-NomeCompleto",
                "Teste Pacote Portatil",
                "-SemPivotNativo",
            ),
            cwd=external_cwd,
            environment=environment,
            timeout=180,
            capture_output=True,
        )

        if hashlib.sha256(input_file.read_bytes()).hexdigest() != original_hash:
            raise RuntimeError("A entrada sintética foi alterada pelo executável.")

        generated: dict[str, list[Path]] = {}
        for directory_name in ("output", "backup", "logs"):
            generated[directory_name] = [
                path
                for path in (smoke_package / directory_name).iterdir()
                if path.is_file()
            ]
            if len(generated[directory_name]) != 1:
                raise RuntimeError(
                    f"Quantidade inesperada de arquivos em {directory_name}: "
                    f"{len(generated[directory_name])}."
                )

        output_workbook = generated["output"][0]
        with ZipFile(output_workbook) as archive:
            names = set(archive.namelist())
        pivot_parts = [
            name for name in names if name.startswith("xl/pivotTables/pivotTable")
        ]
        chart_parts = [name for name in names if name.startswith("xl/charts/chart")]
        if pivot_parts or len(chart_parts) != 1:
            raise RuntimeError(
                "O smoke fallback não produziu a estrutura esperada: "
                f"pivôs={len(pivot_parts)}; gráficos={len(chart_parts)}."
            )

        workbook = load_workbook(output_workbook, data_only=False)
        try:
            if (
                "Apoio_Automacao" not in workbook.sheetnames
                or workbook["Apoio_Automacao"].sheet_state != "hidden"
            ):
                raise RuntimeError("A planilha de apoio não permaneceu oculta.")
        finally:
            workbook.close()


def _zip_info(name: str, *, is_directory: bool) -> ZipInfo:
    info = ZipInfo(name, date_time=ZIP_TIMESTAMP)
    info.compress_type = ZIP_DEFLATED
    info.create_system = 3
    mode = 0o40755 if is_directory else 0o100644
    info.external_attr = mode << 16
    if is_directory:
        info.external_attr |= 0x10
    return info


def create_portable_zip(package_directory: Path, output_file: Path) -> None:
    """Compacta em ordem e com metadados estáveis, incluindo pastas vazias."""

    output_file.parent.mkdir(parents=True, exist_ok=True)
    root_name = package_directory.name
    with ZipFile(
        output_file,
        mode="w",
        compression=ZIP_DEFLATED,
        compresslevel=9,
    ) as archive:
        archive.writestr(_zip_info(f"{root_name}/", is_directory=True), b"")
        for path in sorted(
            package_directory.rglob("*"),
            key=lambda item: item.relative_to(package_directory).as_posix().casefold(),
        ):
            relative = path.relative_to(package_directory).as_posix()
            archive_name = f"{root_name}/{relative}"
            if path.is_dir():
                archive.writestr(
                    _zip_info(f"{archive_name}/", is_directory=True),
                    b"",
                )
            else:
                archive.writestr(
                    _zip_info(archive_name, is_directory=False),
                    path.read_bytes(),
                )


def sha256_file(path: Path) -> str:
    """Calcula SHA-256 em blocos sem carregar o arquivo inteiro."""

    digest = hashlib.sha256()
    with path.open("rb") as file:
        for block in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def write_sha256_file(archive_file: Path) -> Path:
    """Grava o checksum no formato entendido pelas ferramentas usuais."""

    digest = sha256_file(archive_file)
    checksum_file = archive_file.with_name(f"{archive_file.name}.sha256")
    checksum_file.write_text(
        f"{digest}  {archive_file.name}\n",
        encoding="ascii",
        newline="\n",
    )
    return checksum_file


def _parse_arguments(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Constrói o ZIP portátil do Excel Compras Automation.",
    )
    parser.add_argument(
        "--expected-version",
        default=None,
        help="Falha se a versão do projeto não for exatamente esta.",
    )
    parser.add_argument(
        "--commit",
        default=None,
        help="SHA estável registrado no arquivo VERSAO.txt.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    arguments = _parse_arguments(argv)
    project_root = Path(__file__).resolve().parents[1]

    if sys.platform != "win32":
        raise RuntimeError("A distribuição deve ser construída no Windows.")
    if platform.machine().casefold() not in {"amd64", "x86_64"}:
        raise RuntimeError("A distribuição atual exige Windows x64.")

    version = read_project_version(project_root)
    if arguments.expected_version and version != arguments.expected_version:
        raise ValueError(
            f"Versão inesperada: projeto={version}; "
            f"esperada={arguments.expected_version}."
        )

    package_name = portable_package_name(version)
    build_parent = project_root / "build"
    dist_parent = project_root / "dist"
    scratch_directory = build_parent / "portable"
    package_directory = dist_parent / package_name
    archive_file = dist_parent / f"{package_name}.zip"
    checksum_file = archive_file.with_name(f"{archive_file.name}.sha256")

    build_parent.mkdir(exist_ok=True)
    dist_parent.mkdir(exist_ok=True)
    _safe_remove_directory(scratch_directory, expected_parent=build_parent)
    _safe_remove_directory(package_directory, expected_parent=dist_parent)
    _safe_remove_file(archive_file, expected_parent=dist_parent)
    _safe_remove_file(checksum_file, expected_parent=dist_parent)
    scratch_directory.mkdir()

    print(f"Construindo {package_name}...")
    frozen_directory = build_frozen_application(
        project_root,
        scratch_directory,
        version,
    )
    assemble_portable_package(
        project_root=project_root,
        frozen_directory=frozen_directory,
        package_directory=package_directory,
        version=version,
        commit=arguments.commit or os.environ.get("GITHUB_SHA"),
    )
    validate_portable_package(package_directory)

    print("Executando smoke test isolado do pacote...")
    run_portable_smoke_test(package_directory, version)

    print("Gerando ZIP e checksum...")
    create_portable_zip(package_directory, archive_file)
    written_checksum = write_sha256_file(archive_file)
    print(f"Pacote: {archive_file}")
    print(f"SHA-256: {written_checksum}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

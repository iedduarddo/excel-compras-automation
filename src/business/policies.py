"""Leitura das políticas e cálculo dos valores-base das solicitações."""

from __future__ import annotations

from datetime import date, datetime
from numbers import Real

from openpyxl import Workbook
from openpyxl.utils.datetime import from_excel

from src.core.exceptions import ValidationError
from src.core.models import Policy, TravelResult, WorkbookLayout
from src.services.text import normalize_text


def _as_number(value: object, field: str, row: int) -> float:
    if isinstance(value, bool) or not isinstance(value, Real):
        raise ValidationError(
            f"Linha {row}: o campo '{field}' deve ser numérico. Valor: {value!r}"
        )
    return float(value)


def _as_date(value: object, field: str, row: int) -> date | datetime:
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, Real):
        return from_excel(float(value))
    raise ValidationError(
        f"Linha {row}: o campo '{field}' deve ser uma data válida. Valor: {value!r}"
    )


def read_policies(
    workbook: Workbook,
    layout: WorkbookLayout,
) -> dict[str, Policy]:
    """Lê a tabela de políticas independentemente de sua posição."""

    worksheet = workbook[layout.policies.title]
    columns = layout.policies.columns
    policies: dict[str, Policy] = {}

    for row in range(layout.policies.header_row + 1, worksheet.max_row + 1):
        service = worksheet.cell(row, columns["service_type"]).value
        limit_value = worksheet.cell(row, columns["limit_value"]).value
        min_days = worksheet.cell(row, columns["min_lead_days"]).value
        if service is None and limit_value is None and min_days is None:
            continue
        if not service or not isinstance(limit_value, Real) or not isinstance(
            min_days, Real
        ):
            continue

        key = normalize_text(service)
        if key in policies:
            raise ValidationError(
                f"Política duplicada para o tipo de serviço '{service}'."
            )
        policies[key] = Policy(
            service_type=str(service).strip(),
            limit_value=float(limit_value),
            min_lead_days=int(min_days),
            worksheet_row=row,
        )

    if not policies:
        raise ValidationError(
            "Nenhuma política válida de serviço, limite e antecedência foi encontrada."
        )
    return policies


def find_last_data_row(
    workbook: Workbook,
    layout: WorkbookLayout,
) -> int:
    """Encontra a última solicitação usando a coluna de ID como referência."""

    worksheet = workbook[layout.base.title]
    id_column = layout.base.columns["request_id"]
    for row in range(worksheet.max_row, layout.base.header_row, -1):
        value = worksheet.cell(row, id_column).value
        if value is not None and str(value).strip():
            return row
    raise ValidationError("A base de viagens não possui solicitações.")


def analyze_travels(
    workbook: Workbook,
    layout: WorkbookLayout,
    policies: dict[str, Policy],
) -> list[TravelResult]:
    """Calcula os resultados em memória antes de escrever fórmulas."""

    worksheet = workbook[layout.base.title]
    columns = layout.base.columns
    last_row = find_last_data_row(workbook, layout)
    results: list[TravelResult] = []

    for row in range(layout.base.header_row + 1, last_row + 1):
        request_id = worksheet.cell(row, columns["request_id"]).value
        if request_id is None or not str(request_id).strip():
            continue

        request_date = _as_date(
            worksheet.cell(row, columns["request_date"]).value,
            "Data Solicitação",
            row,
        )
        travel_date = _as_date(
            worksheet.cell(row, columns["travel_date"]).value,
            "Data Viagem",
            row,
        )
        quantity = _as_number(
            worksheet.cell(row, columns["quantity"]).value,
            "Quantidade",
            row,
        )
        unit_value = _as_number(
            worksheet.cell(row, columns["unit_value"]).value,
            "Valor Unitário",
            row,
        )
        fees_value = worksheet.cell(row, columns["fees"]).value
        fees = 0.0 if fees_value in (None, "") else _as_number(fees_value, "Taxas", row)

        service_type = str(
            worksheet.cell(row, columns["service_type"]).value or ""
        ).strip()
        policy = policies.get(normalize_text(service_type))
        total_value = round(quantity * unit_value + fees, 2)
        lead_days = (travel_date - request_date).days

        if policy is None:
            policy_limit = 0.0
            min_lead_days = 0
            policy_status = "Revisar"
        else:
            policy_limit = policy.limit_value
            min_lead_days = policy.min_lead_days
            policy_status = (
                "Fora"
                if total_value > policy_limit or lead_days < min_lead_days
                else "OK"
            )

        results.append(
            TravelResult(
                worksheet_row=row,
                request_id=str(request_id).strip(),
                request_date=request_date,
                travel_date=travel_date,
                service_type=service_type,
                supplier=str(
                    worksheet.cell(row, columns["supplier"]).value or ""
                ).strip(),
                cost_center=str(
                    worksheet.cell(row, columns["cost_center"]).value or ""
                ).strip(),
                booking_status=str(
                    worksheet.cell(row, columns["booking_status"]).value or ""
                ).strip(),
                criticality=str(
                    worksheet.cell(row, columns["criticality"]).value or ""
                ).strip(),
                card_status=str(
                    worksheet.cell(row, columns["card_status"]).value or ""
                ).strip(),
                total_value=total_value,
                lead_days=lead_days,
                policy_limit=policy_limit,
                min_lead_days=min_lead_days,
                policy_status=policy_status,
                limit_difference=round(total_value - policy_limit, 2),
            )
        )

    if not results:
        raise ValidationError("Nenhuma solicitação válida foi encontrada na base.")
    return results


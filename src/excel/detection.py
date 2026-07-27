"""Detecção de abas, linhas de cabeçalho e colunas por conteúdo."""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.core.exceptions import DetectionError
from src.core.models import SheetLayout, WorkbookLayout
from src.services.text import normalize_text, text_similarity

BASE_REQUIRED_FIELDS = (
    "request_id",
    "request_date",
    "travel_date",
    "cost_center",
    "service_type",
    "supplier",
    "quantity",
    "unit_value",
    "fees",
    "booking_status",
    "criticality",
    "card_status",
)

POLICY_REQUIRED_FIELDS = ("service_type", "limit_value", "min_lead_days")
RESPONSE_REQUIRED_FIELDS = ("indicator", "answer")


@dataclass(frozen=True)
class _Candidate:
    """Candidato interno a aba e linha de cabeçalho."""

    sheet: Worksheet
    header_row: int
    columns: dict[str, int]
    score: float


class HeaderDetector:
    """Associa cabeçalhos a campos canônicos usando aliases e similaridade."""

    def __init__(self, threshold: float = 0.82) -> None:
        self.threshold = threshold

    def map_header_row(
        self,
        worksheet: Worksheet,
        row_number: int,
        field_aliases: dict[str, list[str]],
    ) -> dict[str, int]:
        """Mapeia uma linha sem depender da posição física das colunas."""

        candidates: list[tuple[float, str, int]] = []
        for column in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row=row_number, column=column).value
            if value is None or not str(value).strip():
                continue
            for canonical, aliases in field_aliases.items():
                score = max(
                    text_similarity(value, alias)
                    for alias in [canonical.replace("_", " "), *aliases]
                )
                if score >= self.threshold:
                    candidates.append((score, canonical, column))

        candidates.sort(key=lambda item: item[0], reverse=True)
        assigned_fields: set[str] = set()
        assigned_columns: set[int] = set()
        mapping: dict[str, int] = {}
        for _, canonical, column in candidates:
            if canonical in assigned_fields or column in assigned_columns:
                continue
            mapping[canonical] = column
            assigned_fields.add(canonical)
            assigned_columns.add(column)
        return mapping

    def best_header(
        self,
        worksheet: Worksheet,
        field_aliases: dict[str, list[str]],
        required_fields: Iterable[str],
        scan_rows: int = 30,
    ) -> tuple[int, dict[str, int], int]:
        """Retorna a linha com mais campos obrigatórios reconhecidos."""

        required = tuple(required_fields)
        best_row = 0
        best_mapping: dict[str, int] = {}
        best_hits = -1
        best_total = -1

        for row in range(1, min(worksheet.max_row, scan_rows) + 1):
            mapping = self.map_header_row(worksheet, row, field_aliases)
            hits = sum(field in mapping for field in required)
            if (hits, len(mapping)) > (best_hits, best_total):
                best_row = row
                best_mapping = mapping
                best_hits = hits
                best_total = len(mapping)

        return best_row, best_mapping, max(best_hits, 0)


class WorkbookDetector:
    """Identifica a estrutura da pasta de trabalho por nomes e conteúdo."""

    def __init__(self, aliases: dict[str, object]) -> None:
        self.aliases = aliases
        self.header_detector = HeaderDetector()

    def detect(self, workbook: Workbook) -> WorkbookLayout:
        """Detecta as três áreas necessárias e valida colisões."""

        base = self._detect_role(
            workbook,
            role="base",
            column_group="base_columns",
            required_fields=BASE_REQUIRED_FIELDS,
            min_hits=10,
        )
        policies = self._detect_role(
            workbook,
            role="policies",
            column_group="policy_columns",
            required_fields=POLICY_REQUIRED_FIELDS,
            min_hits=3,
        )
        responses = self._detect_role(
            workbook,
            role="responses",
            column_group="response_columns",
            required_fields=RESPONSE_REQUIRED_FIELDS,
            min_hits=2,
        )

        titles = {base.title, policies.title, responses.title}
        if len(titles) != 3:
            raise DetectionError(
                "A mesma aba foi classificada em mais de uma função. "
                "Confira os nomes e cabeçalhos das abas."
            )
        return WorkbookLayout(base=base, policies=policies, responses=responses)

    def _detect_role(
        self,
        workbook: Workbook,
        role: str,
        column_group: str,
        required_fields: tuple[str, ...],
        min_hits: int,
    ) -> SheetLayout:
        field_aliases = self.aliases[column_group]
        sheet_aliases = self.aliases["sheets"][role]
        candidates: list[_Candidate] = []

        for worksheet in workbook.worksheets:
            header_row, columns, hits = self.header_detector.best_header(
                worksheet,
                field_aliases=field_aliases,
                required_fields=required_fields,
            )
            name_score = max(
                text_similarity(worksheet.title, alias) for alias in sheet_aliases
            )
            coverage = hits / len(required_fields)
            score = coverage * 100 + name_score * 30 + len(columns) * 0.25
            candidates.append(
                _Candidate(
                    sheet=worksheet,
                    header_row=header_row,
                    columns=columns,
                    score=score,
                )
            )

        candidates.sort(key=lambda item: item.score, reverse=True)
        winner = candidates[0]
        actual_hits = sum(field in winner.columns for field in required_fields)
        missing = [field for field in required_fields if field not in winner.columns]
        if actual_hits < min_hits:
            readable = ", ".join(missing)
            raise DetectionError(
                f"Não foi possível identificar com segurança a aba '{role}'. "
                f"Melhor candidata: '{winner.sheet.title}'. "
                f"Campos não encontrados: {readable}."
            )

        return SheetLayout(
            title=winner.sheet.title,
            header_row=winner.header_row,
            columns=dict(winner.columns),
        )


def find_indicator_rows(
    worksheet: Worksheet,
    aliases: dict[str, list[str]],
    max_rows: int = 80,
    max_columns: int = 12,
) -> dict[str, tuple[int, int]]:
    """Localiza os rótulos dos indicadores, mesmo que tenham mudado de linha."""

    found: dict[str, tuple[int, int]] = {}
    candidates: list[tuple[float, str, int, int]] = []

    for row in range(1, min(worksheet.max_row, max_rows) + 1):
        for column in range(1, min(worksheet.max_column, max_columns) + 1):
            value = worksheet.cell(row=row, column=column).value
            if value is None or not normalize_text(value):
                continue
            for canonical, label_aliases in aliases.items():
                score = max(text_similarity(value, alias) for alias in label_aliases)
                if score >= 0.86:
                    candidates.append((score, canonical, row, column))

    candidates.sort(key=lambda item: item[0], reverse=True)
    occupied_cells: set[tuple[int, int]] = set()
    for _, canonical, row, column in candidates:
        if canonical in found or (row, column) in occupied_cells:
            continue
        found[canonical] = (row, column)
        occupied_cells.add((row, column))

    missing = [canonical for canonical in aliases if canonical not in found]
    if missing:
        raise DetectionError(
            "Os seguintes indicadores não foram encontrados na aba de respostas: "
            + ", ".join(missing)
        )
    return found

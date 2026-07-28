"""Indicadores, prioridades, resumo de compatibilidade e gráfico."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.business.priorities import rank_immediate_requests
from src.core.models import TravelResult, WorkbookLayout
from src.excel._writer_common import (
    DARK_BLUE,
    LIGHT_BLUE,
    LIGHT_YELLOW,
    THIN_GRAY,
    _absolute_range,
    _cell,
    _style_header_range,
    _style_total_range,
)
from src.excel.detection import find_indicator_rows
from src.services.text import normalize_text, text_similarity


def write_responses(
    workbook: Workbook,
    layout: WorkbookLayout,
    aliases: dict[str, object],
    travels: list[TravelResult],
    support_refs: dict[str, str],
    last_row: int,
    top_quantity: int,
) -> int:
    """Preenche indicadores, as cinco prioridades e reserva a área do pivô."""

    worksheet = workbook[layout.responses.title]
    indicator_rows = find_indicator_rows(
        worksheet,
        aliases["indicator_labels"],
    )
    answer_column = layout.responses.columns["answer"]
    formula_column = layout.responses.columns.get("formula_used")
    base_columns = layout.base.columns
    first_row = layout.base.header_row + 1

    ranges = {
        key: _absolute_range(layout.base.title, column, first_row, last_row)
        for key, column in base_columns.items()
    }
    formulas = {
        "total_travel_value": f"=SUM({ranges['total_value']})",
        "outside_policy_value": (
            f'=SUMIFS({ranges["total_value"]},{ranges["policy_status"]},"Fora")'
        ),
        "outside_policy_count": (f'=COUNTIF({ranges["policy_status"]},"Fora")'),
        "outside_policy_percent": (
            f'=IFERROR(COUNTIF({ranges["policy_status"]},"Fora")/'
            f"COUNTA({ranges['request_id']}),0)"
        ),
        "card_issue_value": (
            f'=SUMIFS({ranges["total_value"]},{ranges["card_status"]},"Pendente")'
            f'+SUMIFS({ranges["total_value"]},{ranges["card_status"]},"Divergente")'
        ),
        "emergency_count": (f'=COUNTIF({ranges["criticality"]},"Emergencial")'),
        "top_cost_center": (
            f"=INDEX({support_refs['cost_center_labels']},"
            f"MATCH(MAX({support_refs['cost_center_values']}),"
            f"{support_refs['cost_center_values']},0))"
        ),
        "top_supplier": (
            f"=INDEX({support_refs['supplier_labels']},"
            f"MATCH(MAX({support_refs['supplier_values']}),"
            f"{support_refs['supplier_values']},0))"
        ),
        "potential_savings": (
            f"=SUMIFS({ranges['limit_difference']},"
            f'{ranges["policy_status"]},"Fora",'
            f'{ranges["limit_difference"]},">0")'
        ),
        "average_ticket": f"=AVERAGE({ranges['total_value']})",
    }
    descriptions = {
        "total_travel_value": "SOMA",
        "outside_policy_value": "SOMASES",
        "outside_policy_count": "CONT.SE",
        "outside_policy_percent": "CONT.SE / CONT.VALORES",
        "card_issue_value": "SOMASES + SOMASES",
        "emergency_count": "CONT.SE",
        "top_cost_center": "ÍNDICE / CORRESP / MÁXIMO",
        "top_supplier": "ÍNDICE / CORRESP / MÁXIMO",
        "potential_savings": "SOMASES com diferença positiva",
        "average_ticket": "MÉDIA",
    }
    currency_indicators = {
        "total_travel_value",
        "outside_policy_value",
        "card_issue_value",
        "potential_savings",
        "average_ticket",
    }

    for key, (row, _) in indicator_rows.items():
        answer = worksheet.cell(row, answer_column)
        answer.value = formulas[key]
        if key in currency_indicators:
            answer.number_format = "R$ #,##0.00"
        elif key == "outside_policy_percent":
            answer.number_format = "0.0%"
        elif key in {"outside_policy_count", "emergency_count"}:
            answer.number_format = "0"
        if formula_column:
            worksheet.cell(row, formula_column, descriptions[key])

    priority_header_row, priority_columns = _find_priority_table(worksheet)
    _write_top_requests(
        worksheet=worksheet,
        header_row=priority_header_row,
        columns=priority_columns,
        ranked=rank_immediate_requests(travels, top_quantity),
        quantity=top_quantity,
    )

    pivot_title_row = _find_pivot_title_row(worksheet)
    pivot_start_row = pivot_title_row + 2
    _prepare_pivot_area(worksheet, pivot_title_row, pivot_start_row)
    _format_response_sheet(worksheet, priority_header_row)
    return pivot_start_row


def create_fallback_summary_and_chart(
    workbook: Workbook,
    layout: WorkbookLayout,
    travels: list[TravelResult],
    pivot_start_row: int,
    last_row: int,
) -> None:
    """Cria um resumo formula-driven caso o Excel Desktop não esteja disponível."""

    worksheet = workbook[layout.responses.title]
    columns = layout.base.columns
    first_row = layout.base.header_row + 1
    cost_centers = sorted(
        {item.cost_center for item in travels if item.cost_center},
        key=normalize_text,
    )
    services = sorted(
        {item.service_type for item in travels if item.service_type},
        key=normalize_text,
    )
    total_range = _absolute_range(
        layout.base.title,
        columns["total_value"],
        first_row,
        last_row,
    )
    cost_center_range = _absolute_range(
        layout.base.title,
        columns["cost_center"],
        first_row,
        last_row,
    )
    service_range = _absolute_range(
        layout.base.title,
        columns["service_type"],
        first_row,
        last_row,
    )

    worksheet.cell(pivot_start_row, 1, "Centro de Custo")
    for index, service in enumerate(services, start=2):
        worksheet.cell(pivot_start_row, index, service)
    total_column = 2 + len(services)
    worksheet.cell(pivot_start_row, total_column, "Total Geral")
    _style_header_range(
        worksheet,
        f"A{pivot_start_row}:{get_column_letter(total_column)}{pivot_start_row}",
    )

    for row_offset, cost_center in enumerate(cost_centers, start=1):
        row = pivot_start_row + row_offset
        worksheet.cell(row, 1, cost_center)
        for column_offset, _ in enumerate(services, start=2):
            service_header = _cell(column_offset, pivot_start_row)
            worksheet.cell(
                row,
                column_offset,
                f"=SUMIFS({total_range},{cost_center_range},$A{row},"
                f"{service_range},{service_header})",
            )
            worksheet.cell(row, column_offset).number_format = "R$ #,##0.00"
        worksheet.cell(
            row,
            total_column,
            f"=SUM(B{row}:{get_column_letter(total_column - 1)}{row})",
        )
        worksheet.cell(row, total_column).number_format = "R$ #,##0.00"

    total_row = pivot_start_row + len(cost_centers) + 1
    worksheet.cell(total_row, 1, "Total Geral")
    for column in range(2, total_column + 1):
        letter = get_column_letter(column)
        worksheet.cell(
            total_row,
            column,
            f"=SUM({letter}{pivot_start_row + 1}:{letter}{total_row - 1})",
        )
        worksheet.cell(total_row, column).number_format = "R$ #,##0.00"
    _style_total_range(
        worksheet,
        f"A{total_row}:{get_column_letter(total_column)}{total_row}",
    )

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.grouping = "clustered"
    chart.title = "Valor Total por Centro de Custo e Tipo de Serviço"
    chart.y_axis.title = "Valor Total (R$)"
    chart.x_axis.title = "Centro de Custo"
    chart.height = 9
    chart.width = 18
    data = Reference(
        worksheet,
        min_col=2,
        max_col=total_column - 1,
        min_row=pivot_start_row,
        max_row=pivot_start_row + len(cost_centers),
    )
    categories = Reference(
        worksheet,
        min_col=1,
        min_row=pivot_start_row + 1,
        max_row=pivot_start_row + len(cost_centers),
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend.position = "b"
    worksheet.add_chart(
        chart, f"{get_column_letter(total_column + 2)}{pivot_start_row}"
    )


def _find_priority_table(worksheet: Worksheet) -> tuple[int, dict[str, int]]:
    aliases = {
        "request_id": ["ID Solicitação", "ID da Solicitação"],
        "reason": ["Motivo da prioridade", "Justificativa"],
        "action": ["Ação recomendada", "Ação"],
        "order": ["Ordem (1 a 5)", "Ordem", "Ranking"],
    }
    for row in range(1, min(worksheet.max_row, 80) + 1):
        mapping: dict[str, int] = {}
        for column in range(1, min(worksheet.max_column, 12) + 1):
            value = worksheet.cell(row, column).value
            for canonical, options in aliases.items():
                if (
                    value is not None
                    and max(text_similarity(value, option) for option in options)
                    >= 0.86
                ):
                    mapping[canonical] = column
        if len(mapping) >= 3 and "request_id" in mapping:
            defaults = {
                "request_id": mapping["request_id"],
                "reason": mapping.get("reason", mapping["request_id"] + 1),
                "action": mapping.get("action", mapping["request_id"] + 2),
                "order": mapping.get("order", mapping["request_id"] + 3),
            }
            return row, defaults

    row = worksheet.max_row + 2
    return row, {"request_id": 1, "reason": 2, "action": 3, "order": 4}


def _write_top_requests(
    worksheet: Worksheet,
    header_row: int,
    columns: dict[str, int],
    ranked: list[TravelResult],
    quantity: int,
) -> None:
    headers = {
        "request_id": "ID Solicitação",
        "reason": "Motivo da prioridade",
        "action": "Ação recomendada",
        "order": "Ordem (1 a 5)",
    }
    for canonical, title in headers.items():
        worksheet.cell(header_row, columns[canonical], title)
    first_column = min(columns.values())
    last_column = max(columns.values())
    _style_header_range(
        worksheet,
        f"{_cell(first_column, header_row)}:{_cell(last_column, header_row)}",
    )

    for offset in range(1, quantity + 1):
        row = header_row + offset
        for column in range(first_column, last_column + 1):
            cell = worksheet.cell(row, column)
            cell.value = None
            cell.fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
            cell.border = Border(bottom=THIN_GRAY)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for order, travel in enumerate(ranked, start=1):
        row = header_row + order
        reasons = "; ".join(travel.reasons[:4]) or "maior score combinado de risco"
        actions = "; ".join(travel.recommended_actions[:3]) or "revisar a solicitação"
        worksheet.cell(row, columns["request_id"], travel.request_id)
        worksheet.cell(row, columns["reason"], reasons[:1].upper() + reasons[1:] + ".")
        worksheet.cell(row, columns["action"], actions[:1].upper() + actions[1:] + ".")
        order_cell = worksheet.cell(row, columns["order"], order)
        order_cell.comment = Comment(
            f"Score de prioridade calculado: {travel.score:.2f}",
            "Excel Compras Automation",
        )
        # Altura fixa generosa porque o Excel não oferece AutoFit confiável
        # para células preenchidas por openpyxl antes da abertura do arquivo.
        worksheet.row_dimensions[row].height = 115


def _find_pivot_title_row(worksheet: Worksheet) -> int:
    aliases = [
        "Tabela dinâmica e gráfico",
        "Tabela dinamica e grafico",
        "Resumo por centro de custo",
    ]
    for row in range(1, min(worksheet.max_row, 100) + 1):
        for column in range(1, min(worksheet.max_column, 12) + 1):
            value = worksheet.cell(row, column).value
            if (
                value is not None
                and max(text_similarity(value, alias) for alias in aliases) >= 0.86
            ):
                return row
    return worksheet.max_row + 2


def _prepare_pivot_area(
    worksheet: Worksheet,
    title_row: int,
    pivot_start_row: int,
) -> None:
    for merged in list(worksheet.merged_cells.ranges):
        if merged.min_row > title_row:
            worksheet.unmerge_cells(str(merged))

    worksheet.merge_cells(
        start_row=title_row,
        start_column=1,
        end_row=title_row,
        end_column=8,
    )
    worksheet.cell(title_row, 1, "Tabela dinâmica e gráfico")
    for cell in worksheet[title_row][0:8]:
        cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cell.font = Font(color=DARK_BLUE, bold=True)

    for row in range(title_row + 1, min(worksheet.max_row + 20, title_row + 45)):
        for column in range(1, 16):
            worksheet.cell(row, column).value = None
    worksheet.cell(
        title_row + 1,
        1,
        "Valor Total por Centro de Custo e Tipo de Serviço",
    )
    worksheet.merge_cells(
        start_row=title_row + 1,
        start_column=1,
        end_row=title_row + 1,
        end_column=6,
    )
    worksheet.cell(title_row + 1, 1).font = Font(italic=True, color="FF7F6000")
    worksheet.cell(title_row + 1, 1).alignment = Alignment(
        vertical="center",
        wrap_text=True,
    )
    worksheet.row_dimensions[pivot_start_row - 1].height = 30


def _format_response_sheet(worksheet: Worksheet, priority_header_row: int) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.column_dimensions["A"].width = 42
    worksheet.column_dimensions["B"].width = 33
    worksheet.column_dimensions["C"].width = 45
    worksheet.column_dimensions["D"].width = 16
    for row in range(1, max(worksheet.max_row, priority_header_row + 5) + 1):
        for column in range(1, 5):
            worksheet.cell(row, column).alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )

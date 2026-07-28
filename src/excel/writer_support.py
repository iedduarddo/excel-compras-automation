"""Aba de apoio auditável usada por fórmulas, resumos e PivotTable."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.core.models import TravelResult, WorkbookLayout
from src.excel._writer_common import (
    DARK_BLUE,
    WHITE,
    _absolute_range,
    _quoted_sheet,
    _style_header_range,
)
from src.services.text import normalize_text

SUPPORT_SHEET = "Apoio_Automacao"


def create_support_sheet(
    workbook: Workbook,
    layout: WorkbookLayout,
    travels: list[TravelResult],
    rules: dict[str, object],
    last_row: int,
) -> dict[str, str]:
    """Cria premissas visíveis às fórmulas e resumos auxiliares."""

    if SUPPORT_SHEET in workbook.sheetnames:
        del workbook[SUPPORT_SHEET]
    worksheet = workbook.create_sheet(SUPPORT_SHEET)
    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells("A1:I1")
    worksheet["A1"] = "APOIO DA AUTOMAÇÃO — REGRAS E RESUMOS AUDITÁVEIS"
    worksheet["A1"].fill = PatternFill("solid", fgColor=DARK_BLUE)
    worksheet["A1"].font = Font(color=WHITE, bold=True, size=14)
    worksheet["A1"].alignment = Alignment(horizontal="center")

    worksheet["A3"] = "Regra"
    worksheet["B3"] = "Valor"
    worksheet["C3"] = "Aplicação"
    _style_header_range(worksheet, "A3:C3")

    weights = rules["priority_weights"]
    thresholds = rules["priority_thresholds"]
    rule_rows = [
        ("card_divergent", "Cartão divergente", weights["card_divergent"]),
        ("card_pending", "Cartão pendente", weights["card_pending"]),
        ("card_other_issue", "Outro status de cartão", weights["card_other_issue"]),
        (
            "criticality_emergency",
            "Criticidade emergencial",
            weights["criticality_emergency"],
        ),
        (
            "criticality_executive",
            "Criticidade executiva",
            weights["criticality_executive"],
        ),
        ("outside_policy", "Fora da política", weights["outside_policy"]),
        (
            "policy_not_found",
            "Política não localizada",
            weights["policy_not_found"],
        ),
        ("booking_pending", "Reserva pendente", weights["booking_pending"]),
        (
            "booking_reschedule",
            "Reserva em remarcação",
            weights["booking_reschedule"],
        ),
        (
            "cost_over_limit_max",
            "Máximo por excesso de custo",
            weights["cost_over_limit_max"],
        ),
        (
            "lead_time_shortfall_max",
            "Máximo por falta de antecedência",
            weights["lead_time_shortfall_max"],
        ),
        (
            "total_value_max",
            "Máximo por valor total",
            weights["total_value_max"],
        ),
        ("critical_threshold", "Limite — prioridade crítica", thresholds["critical"]),
        ("high_threshold", "Limite — prioridade alta", thresholds["high"]),
    ]
    rule_cells: dict[str, str] = {}
    for row, (key, label, value) in enumerate(rule_rows, start=4):
        worksheet.cell(row, 1, label)
        worksheet.cell(row, 2, value)
        worksheet.cell(row, 3, "Componente do score de prioridade")
        rule_cells[key] = f"'{SUPPORT_SHEET}'!$B${row}"

    base_sheet = _quoted_sheet(layout.base.title)
    base_columns = layout.base.columns
    first_data_row = layout.base.header_row + 1
    total_range = _absolute_range(
        layout.base.title,
        base_columns["total_value"],
        first_data_row,
        last_row,
    )
    cost_center_range = _absolute_range(
        layout.base.title,
        base_columns["cost_center"],
        first_data_row,
        last_row,
    )
    supplier_range = _absolute_range(
        layout.base.title,
        base_columns["supplier"],
        first_data_row,
        last_row,
    )

    worksheet["E3"] = "Centro de Custo"
    worksheet["F3"] = "Valor Total"
    _style_header_range(worksheet, "E3:F3")
    cost_centers = sorted(
        {item.cost_center for item in travels if item.cost_center},
        key=normalize_text,
    )
    for row, cost_center in enumerate(cost_centers, start=4):
        worksheet.cell(row, 5, cost_center)
        worksheet.cell(
            row,
            6,
            f"=SUMIF({cost_center_range},E{row},{total_range})",
        )
        worksheet.cell(row, 6).number_format = "R$ #,##0.00"

    worksheet["H3"] = "Fornecedor"
    worksheet["I3"] = "Valor Total"
    _style_header_range(worksheet, "H3:I3")
    suppliers = sorted(
        {item.supplier for item in travels if item.supplier},
        key=normalize_text,
    )
    for row, supplier in enumerate(suppliers, start=4):
        worksheet.cell(row, 8, supplier)
        worksheet.cell(
            row,
            9,
            f"=SUMIF({supplier_range},H{row},{total_range})",
        )
        worksheet.cell(row, 9).number_format = "R$ #,##0.00"

    for column, width in {
        "A": 30,
        "B": 12,
        "C": 35,
        "E": 20,
        "F": 16,
        "H": 24,
        "I": 16,
    }.items():
        worksheet.column_dimensions[column].width = width
    for row in worksheet.iter_rows(min_row=4, max_row=max(17, 3 + len(suppliers))):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Fonte estática e auditável para a Tabela Dinâmica nativa. Os valores
    # vêm do cálculo Python já validado, enquanto as fórmulas permanecem na
    # Base_Viagens para atender ao teste e permitir auditoria.
    worksheet["K3"] = "Centro de Custo"
    worksheet["L3"] = "Tipo de Serviço"
    worksheet["M3"] = "Valor Total"
    _style_header_range(worksheet, "K3:M3")
    for row, travel in enumerate(travels, start=4):
        worksheet.cell(row, 11, travel.cost_center)
        worksheet.cell(row, 12, travel.service_type)
        worksheet.cell(row, 13, travel.total_value)
        worksheet.cell(row, 13).number_format = "R$ #,##0.00"
    worksheet.column_dimensions["K"].width = 20
    worksheet.column_dimensions["L"].width = 24
    worksheet.column_dimensions["M"].width = 16

    worksheet.sheet_state = "hidden"
    return {
        **rule_cells,
        "cost_center_labels": (f"'{SUPPORT_SHEET}'!$E$4:$E${3 + len(cost_centers)}"),
        "cost_center_values": (f"'{SUPPORT_SHEET}'!$F$4:$F${3 + len(cost_centers)}"),
        "supplier_labels": f"'{SUPPORT_SHEET}'!$H$4:$H${3 + len(suppliers)}",
        "supplier_values": f"'{SUPPORT_SHEET}'!$I$4:$I${3 + len(suppliers)}",
        "base_sheet": base_sheet,
    }

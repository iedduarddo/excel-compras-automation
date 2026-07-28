"""Constantes e helpers compartilhados pelos escritores de planilha."""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

BLUE = "FF2F75B5"
DARK_BLUE = "FF17365D"
ORANGE = "FFC55A11"
LIGHT_YELLOW = "FFFFF2CC"
LIGHT_BLUE = "FFD9EAF7"
LIGHT_RED = "FFF4CCCC"
LIGHT_ORANGE = "FFFCE4D6"
DARK_RED = "FF9C0006"
WHITE = "FFFFFFFF"
GRAY = "FFD9E1F2"
THIN_GRAY = Side(style="thin", color="FFD9E1F2")


def _quoted_sheet(title: str) -> str:
    return "'" + title.replace("'", "''") + "'"


def _absolute_range(title: str, column: int, first_row: int, last_row: int) -> str:
    letter = get_column_letter(column)
    return f"{_quoted_sheet(title)}!${letter}${first_row}:${letter}${last_row}"


def _cell(column: int, row: int) -> str:
    return f"{get_column_letter(column)}{row}"


def _style_header_range(worksheet: Worksheet, cell_range: str) -> None:
    for row in worksheet[cell_range]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.font = Font(color=WHITE, bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = Border(bottom=THIN_GRAY)


def _style_total_range(worksheet: Worksheet, cell_range: str) -> None:
    for row in worksheet[cell_range]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            cell.font = Font(color=DARK_BLUE, bold=True)
            cell.border = Border(top=THIN_GRAY)

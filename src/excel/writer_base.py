"""Estrutura, fórmulas e formatação da base de viagens."""

from __future__ import annotations

from copy import copy
from dataclasses import replace

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from src.core.models import Policy, WorkbookLayout
from src.excel._writer_common import (
    DARK_BLUE,
    DARK_RED,
    GRAY,
    LIGHT_ORANGE,
    LIGHT_RED,
    LIGHT_YELLOW,
    ORANGE,
    WHITE,
    _absolute_range,
    _cell,
)

DERIVED_HEADERS = {
    "total_value": "Valor Total",
    "lead_days": "Dias Antecedência",
    "policy_limit": "Limite Política",
    "policy_status": "Status Política",
    "limit_difference": "Diferença p/ Limite",
    "priority": "Prioridade",
    "priority_score": "Score Prioridade",
}


def ensure_derived_columns(
    workbook: Workbook,
    layout: WorkbookLayout,
    last_row: int,
) -> WorkbookLayout:
    """Localiza ou adiciona as colunas calculadas e a coluna auxiliar."""

    worksheet = workbook[layout.base.title]
    columns = dict(layout.base.columns)
    next_column = max(worksheet.max_column, *columns.values()) + 1

    for canonical, header in DERIVED_HEADERS.items():
        if canonical in columns:
            continue
        column = next_column
        next_column += 1
        columns[canonical] = column
        header_cell = worksheet.cell(layout.base.header_row, column, header)
        _style_new_header(header_cell, auxiliary=canonical == "priority_score")

        source_column = max(1, column - 1)
        for row in range(layout.base.header_row + 1, last_row + 1):
            source = worksheet.cell(row, source_column)
            target = worksheet.cell(row, column)
            target._style = copy(source._style)
            target.fill = PatternFill(
                "solid",
                fgColor=GRAY if canonical == "priority_score" else LIGHT_YELLOW,
            )

    worksheet.column_dimensions[
        get_column_letter(columns["priority_score"])
    ].hidden = True
    worksheet.column_dimensions[get_column_letter(columns["total_value"])].width = 15
    worksheet.column_dimensions[get_column_letter(columns["lead_days"])].width = 13
    worksheet.column_dimensions[get_column_letter(columns["policy_limit"])].width = 15
    worksheet.column_dimensions[get_column_letter(columns["policy_status"])].width = 15
    worksheet.column_dimensions[
        get_column_letter(columns["limit_difference"])
    ].width = 18
    worksheet.column_dimensions[get_column_letter(columns["priority"])].width = 13

    _extend_or_create_table(
        worksheet=worksheet,
        header_row=layout.base.header_row,
        last_row=last_row,
        last_column=max(columns.values()),
    )

    updated_base = replace(layout.base, columns=columns)
    return replace(layout, base=updated_base)


def write_base_formulas(
    workbook: Workbook,
    layout: WorkbookLayout,
    policies: dict[str, Policy],
    support_refs: dict[str, str],
    last_row: int,
) -> None:
    """Preenche todas as colunas calculadas com fórmulas auditáveis."""

    worksheet = workbook[layout.base.title]
    columns = layout.base.columns
    first_row = layout.base.header_row + 1
    policy_rows = sorted(policy.worksheet_row for policy in policies.values())
    policy_first = min(policy_rows)
    policy_last = max(policy_rows)
    policy_sheet = layout.policies.title
    policy_columns = layout.policies.columns

    service_policy_range = _absolute_range(
        policy_sheet,
        policy_columns["service_type"],
        policy_first,
        policy_last,
    )
    limit_policy_range = _absolute_range(
        policy_sheet,
        policy_columns["limit_value"],
        policy_first,
        policy_last,
    )
    days_policy_range = _absolute_range(
        policy_sheet,
        policy_columns["min_lead_days"],
        policy_first,
        policy_last,
    )

    for row in range(first_row, last_row + 1):
        reference = {
            key: _cell(columns[key], row)
            for key in (
                "request_date",
                "travel_date",
                "service_type",
                "quantity",
                "unit_value",
                "fees",
                "booking_status",
                "criticality",
                "card_status",
                "total_value",
                "lead_days",
                "policy_limit",
                "policy_status",
                "limit_difference",
                "priority",
                "priority_score",
            )
        }
        min_days_lookup = (
            f"IFERROR(INDEX({days_policy_range},"
            f"MATCH({reference['service_type']},{service_policy_range},0)),0)"
        )

        worksheet[reference["total_value"]] = (
            f"=IFERROR({reference['quantity']}*{reference['unit_value']}"
            f"+{reference['fees']},0)"
        )
        worksheet[reference["lead_days"]] = (
            f"=IFERROR({reference['travel_date']}-{reference['request_date']},0)"
        )
        worksheet[reference["policy_limit"]] = (
            f"=IFERROR(INDEX({limit_policy_range},"
            f"MATCH({reference['service_type']},{service_policy_range},0)),0)"
        )
        worksheet[reference["policy_status"]] = (
            f'=IF({reference["policy_limit"]}=0,"Revisar",'
            f"IF(OR({reference['total_value']}>{reference['policy_limit']},"
            f'{reference["lead_days"]}<{min_days_lookup}),"Fora","OK"))'
        )
        worksheet[reference["limit_difference"]] = (
            f"=IFERROR({reference['total_value']}-{reference['policy_limit']},0)"
        )

        card_score = (
            f'IF({reference["card_status"]}="Divergente",'
            f"{support_refs['card_divergent']},"
            f'IF({reference["card_status"]}="Pendente",'
            f"{support_refs['card_pending']},"
            f'IF(AND({reference["card_status"]}<>"",'
            f'{reference["card_status"]}<>"Conferido"),'
            f"{support_refs['card_other_issue']},0)))"
        )
        criticality_score = (
            f'IF({reference["criticality"]}="Emergencial",'
            f"{support_refs['criticality_emergency']},"
            f'IF({reference["criticality"]}="Executivo",'
            f"{support_refs['criticality_executive']},0))"
        )
        policy_score = (
            f'IF({reference["policy_status"]}="Revisar",'
            f"{support_refs['policy_not_found']},"
            f'IF({reference["policy_status"]}="Fora",'
            f"{support_refs['outside_policy']},0))"
        )
        booking_score = (
            f'IF({reference["booking_status"]}="Pendente",'
            f"{support_refs['booking_pending']},"
            f'IF({reference["booking_status"]}="Remarcação",'
            f"{support_refs['booking_reschedule']},0))"
        )
        cost_score = (
            f"IF(AND({reference['policy_limit']}>0,"
            f"{reference['limit_difference']}>0),"
            f"MIN({support_refs['cost_over_limit_max']},"
            f"({reference['limit_difference']}/{reference['policy_limit']})"
            f"*{support_refs['cost_over_limit_max']}),0)"
        )
        lead_score = (
            f"IFERROR(IF({reference['lead_days']}<{min_days_lookup},"
            f"MIN({support_refs['lead_time_shortfall_max']},"
            f"(({min_days_lookup}-{reference['lead_days']})/{min_days_lookup})"
            f"*{support_refs['lead_time_shortfall_max']}),0),0)"
        )
        total_range = (
            f"${get_column_letter(columns['total_value'])}${first_row}:"
            f"${get_column_letter(columns['total_value'])}${last_row}"
        )
        value_score = (
            f"IFERROR(({reference['total_value']}/MAX({total_range}))"
            f"*{support_refs['total_value_max']},0)"
        )
        worksheet[reference["priority_score"]] = (
            f"=ROUND({card_score}+{criticality_score}+{policy_score}"
            f"+{booking_score}+{cost_score}+{lead_score}+{value_score},2)"
        )
        worksheet[reference["priority"]] = (
            f"=IF({reference['priority_score']}>={support_refs['critical_threshold']},"
            f'"Crítica",IF({reference["priority_score"]}>='
            f'{support_refs["high_threshold"]},"Alta","Normal"))'
        )

    currency_columns = ("total_value", "policy_limit", "limit_difference")
    for canonical in currency_columns:
        letter = get_column_letter(columns[canonical])
        for row in range(first_row, last_row + 1):
            worksheet[f"{letter}{row}"].number_format = "R$ #,##0.00"
    score_letter = get_column_letter(columns["priority_score"])
    for row in range(first_row, last_row + 1):
        worksheet[f"{score_letter}{row}"].number_format = "0.00"

    worksheet.cell(
        layout.base.header_row,
        columns["priority"],
    ).comment = Comment(
        "Classificação derivada do Score Prioridade. O score considera custo, "
        "antecedência, cartão corporativo, criticidade, reserva e aderência à política.",
        "Excel Compras Automation",
    )


def apply_conditional_formatting(
    workbook: Workbook,
    layout: WorkbookLayout,
    last_row: int,
) -> None:
    """Destaca linhas fora da política e solicitações críticas."""

    worksheet = workbook[layout.base.title]
    first_row = layout.base.header_row + 1
    last_visible_column = layout.base.columns["priority"]
    visible_range = f"A{first_row}:{get_column_letter(last_visible_column)}{last_row}"
    policy_letter = get_column_letter(layout.base.columns["policy_status"])
    priority_letter = get_column_letter(layout.base.columns["priority"])

    critical_rule = FormulaRule(
        formula=[f'${priority_letter}{first_row}="Crítica"'],
        stopIfTrue=True,
        fill=PatternFill("solid", fgColor=LIGHT_RED),
        font=Font(color=DARK_RED, bold=True),
    )
    outside_rule = FormulaRule(
        formula=[f'${policy_letter}{first_row}="Fora"'],
        fill=PatternFill("solid", fgColor=LIGHT_ORANGE),
        font=Font(color="FF9C5700", bold=True),
    )
    worksheet.conditional_formatting.add(visible_range, critical_rule)
    outside_range = (
        f"{policy_letter}{first_row}:"
        f"{get_column_letter(layout.base.columns['limit_difference'])}{last_row}"
    )
    worksheet.conditional_formatting.add(outside_range, outside_rule)
    worksheet.freeze_panes = _cell(1, first_row)


def _style_new_header(cell, auxiliary: bool) -> None:
    cell.fill = PatternFill("solid", fgColor=GRAY if auxiliary else ORANGE)
    cell.font = Font(color=WHITE if not auxiliary else DARK_BLUE, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _extend_or_create_table(
    worksheet: Worksheet,
    header_row: int,
    last_row: int,
    last_column: int,
) -> None:
    matching_tables = [
        table
        for table in worksheet.tables.values()
        if table.ref.split(":")[0].rstrip("0123456789")
    ]
    if matching_tables:
        table = matching_tables[0]
        _, _, old_last_column, _ = range_boundaries(table.ref)
        for column in range(old_last_column + 1, last_column + 1):
            table.tableColumns.append(
                TableColumn(
                    id=len(table.tableColumns) + 1,
                    name=str(worksheet.cell(header_row, column).value),
                )
            )
        start = table.ref.split(":")[0]
        end = f"{get_column_letter(last_column)}{last_row}"
        table.ref = f"{start}:{end}"
        return

    table = Table(
        displayName="BaseViagensTable",
        ref=(f"A{header_row}:{get_column_letter(last_column)}{last_row}"),
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)

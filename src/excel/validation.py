"""Validações estruturais e numéricas do arquivo final."""

from __future__ import annotations

from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook

from src.core.exceptions import ValidationError
from src.core.models import TravelResult, WorkbookLayout
from src.excel.detection import find_indicator_rows

FORMULA_ERRORS = {
    "#REF!",
    "#DIV/0!",
    "#VALUE!",
    "#NAME?",
    "#N/A",
    "#NUM!",
    "#NULL!",
}


def validate_output(
    output_file: Path,
    layout: WorkbookLayout,
    aliases: dict[str, object],
    travels: list[TravelResult],
    last_row: int,
    native_pivot_expected: bool,
    cached_values_expected: bool,
) -> dict[str, object]:
    """Confere fórmulas, valores calculados, gráfico, pivô e formatação."""

    formula_workbook = load_workbook(
        output_file,
        data_only=False,
        read_only=False,
        keep_vba=output_file.suffix.casefold() == ".xlsm",
    )
    value_workbook = load_workbook(
        output_file,
        data_only=True,
        read_only=False,
        keep_vba=output_file.suffix.casefold() == ".xlsm",
    )

    base_formula = formula_workbook[layout.base.title]
    base_values = value_workbook[layout.base.title]
    response_formula = formula_workbook[layout.responses.title]
    response_values = value_workbook[layout.responses.title]
    first_row = layout.base.header_row + 1

    derived_fields = (
        "total_value",
        "lead_days",
        "policy_limit",
        "policy_status",
        "limit_difference",
        "priority",
        "priority_score",
    )
    missing_formulas: list[str] = []
    for field in derived_fields:
        column = layout.base.columns[field]
        for row in range(first_row, last_row + 1):
            if base_formula.cell(row, column).data_type != "f":
                missing_formulas.append(base_formula.cell(row, column).coordinate)

    if missing_formulas:
        preview = ", ".join(missing_formulas[:10])
        raise ValidationError(
            "Há células calculadas sem fórmula no arquivo final: " + preview
        )

    indicator_rows = find_indicator_rows(
        response_formula,
        aliases["indicator_labels"],
    )
    answer_column = layout.responses.columns["answer"]
    for key, (row, _) in indicator_rows.items():
        if response_formula.cell(row, answer_column).data_type != "f":
            raise ValidationError(
                f"O indicador '{key}' não contém fórmula no arquivo final."
            )

    errors: list[str] = []
    for worksheet in value_workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() in FORMULA_ERRORS:
                    errors.append(f"{worksheet.title}!{cell.coordinate}={cell.value}")
    if errors:
        raise ValidationError(
            "Foram encontrados erros de fórmula no arquivo final: "
            + ", ".join(errors[:12])
        )

    expected_total = round(sum(item.total_value for item in travels), 2)
    calculated_total: float | None = None
    if cached_values_expected:
        calculated_total = round(
            sum(
                float(
                    base_values.cell(
                        row,
                        layout.base.columns["total_value"],
                    ).value
                    or 0
                )
                for row in range(first_row, last_row + 1)
            ),
            2,
        )
        if abs(expected_total - calculated_total) > 0.01:
            raise ValidationError(
                "O total calculado no Excel não confere com a validação em Python: "
                f"Excel={calculated_total:.2f}; Python={expected_total:.2f}."
            )

        total_indicator_row = indicator_rows["total_travel_value"][0]
        total_indicator = response_values.cell(total_indicator_row, answer_column).value
        if (
            total_indicator is None
            or abs(float(total_indicator) - expected_total) > 0.01
        ):
            raise ValidationError(
                "O indicador de valor total não confere com a base de viagens."
            )

    with ZipFile(output_file) as archive:
        names = set(archive.namelist())
    pivot_parts = sorted(
        name for name in names if name.startswith("xl/pivotTables/pivotTable")
    )
    chart_parts = sorted(name for name in names if name.startswith("xl/charts/chart"))
    if native_pivot_expected and not pivot_parts:
        raise ValidationError(
            "A Tabela Dinâmica nativa não foi encontrada dentro do arquivo .xlsx."
        )
    if not chart_parts:
        raise ValidationError("Nenhum gráfico foi encontrado no arquivo final.")

    conditional_rules = sum(
        len(rules) for rules in base_formula.conditional_formatting._cf_rules.values()
    )
    if conditional_rules < 2:
        raise ValidationError(
            "As regras de formatação condicional não foram encontradas."
        )

    return {
        "travel_rows": len(travels),
        "formula_cells_checked": len(derived_fields) * len(travels),
        "indicator_formulas_checked": len(indicator_rows),
        "total_value": calculated_total,
        "formula_errors": 0,
        "conditional_formatting_rules": conditional_rules,
        "pivot_parts": len(pivot_parts),
        "chart_parts": len(chart_parts),
        "support_sheet_hidden": (
            formula_workbook["Apoio_Automacao"].sheet_state == "hidden"
        ),
    }

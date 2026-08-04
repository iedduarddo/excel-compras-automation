"""Exportação determinística da base de viagens para integração com ERP."""

from __future__ import annotations

import csv
import hashlib
import json
import logging
from dataclasses import dataclass
from datetime import date, datetime
from numbers import Real
from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from src.business.policies import analyze_travels, read_policies
from src.business.priorities import apply_priority_scores
from src.core.exceptions import AutomationError, ValidationError
from src.core.models import TravelResult, WorkbookLayout
from src.excel.detection import WorkbookDetector
from src.settings import load_aliases, load_rules

logger = logging.getLogger(__name__)

ERP_BASENAME = "erp_carga_compras"
ERP_FIELDS = (
    "id_solicitacao",
    "data_solicitacao",
    "data_viagem",
    "viajante",
    "area",
    "c1_cc",
    "tipo_servico",
    "destino",
    "fornecedor",
    "c1_quant",
    "c7_preco",
    "taxas",
    "c7_total",
    "status_reserva",
    "criticidade",
    "status_politica",
    "prioridade",
)


@dataclass(frozen=True)
class ERPExportResult:
    """Artefatos produzidos por uma exportação ERP concluída."""

    json_file: Path
    csv_file: Path
    checksum_file: Path
    record_count: int
    sha256: str


class ERPExporter:
    """Converte uma pasta de trabalho validada em artefatos de carga ERP."""

    def __init__(self, excel_path: str | Path, output_dir: str | Path) -> None:
        self.excel_path = Path(excel_path)
        self.output_dir = Path(output_dir)

    def process_and_export(self) -> ERPExportResult:
        """Valida a planilha, normaliza os registros e grava JSON, CSV e SHA-256."""

        workbook = self._load_workbook()
        try:
            aliases = load_aliases()
            rules = load_rules()
            layout = WorkbookDetector(aliases).detect(workbook)
            policies = read_policies(workbook, layout)
            travels = analyze_travels(workbook, layout, policies)
            apply_priority_scores(travels, rules)
            records = self._build_records(workbook, layout, travels)
        finally:
            workbook.close()

        return self._write_artifacts(records)

    def _load_workbook(self) -> Workbook:
        if not self.excel_path.exists() or not self.excel_path.is_file():
            raise AutomationError(
                f"Planilha para exportação ERP não encontrada: {self.excel_path}"
            )
        if self.excel_path.suffix.casefold() not in {".xlsx", ".xlsm"}:
            raise AutomationError(
                "A exportação ERP exige uma planilha .xlsx ou .xlsm. "
                f"Recebido: {self.excel_path.name}"
            )

        try:
            return load_workbook(
                self.excel_path,
                data_only=False,
                read_only=False,
                keep_vba=self.excel_path.suffix.casefold() == ".xlsm",
            )
        except (BadZipFile, InvalidFileException, OSError) as error:
            raise AutomationError(
                f"Não foi possível abrir a planilha para exportação ERP: "
                f"{self.excel_path}"
            ) from error

    @staticmethod
    def _build_records(
        workbook: Workbook,
        layout: WorkbookLayout,
        travels: list[TravelResult],
    ) -> list[dict[str, str | float]]:
        worksheet = workbook[layout.base.title]
        columns = layout.base.columns
        records: list[dict[str, str | float]] = []

        for travel in travels:
            row = travel.worksheet_row
            records.append(
                {
                    "id_solicitacao": travel.request_id,
                    "data_solicitacao": _iso_date(travel.request_date),
                    "data_viagem": _iso_date(travel.travel_date),
                    "viajante": _cell_text(worksheet, columns, row, "traveler"),
                    "area": _cell_text(worksheet, columns, row, "area"),
                    "c1_cc": travel.cost_center,
                    "tipo_servico": travel.service_type,
                    "destino": _cell_text(
                        worksheet,
                        columns,
                        row,
                        "destination",
                    ),
                    "fornecedor": travel.supplier,
                    "c1_quant": _cell_number(
                        worksheet,
                        columns,
                        row,
                        "quantity",
                        "Quantidade",
                    ),
                    "c7_preco": _cell_number(
                        worksheet,
                        columns,
                        row,
                        "unit_value",
                        "Valor Unitário",
                    ),
                    "taxas": _cell_number(
                        worksheet,
                        columns,
                        row,
                        "fees",
                        "Taxas",
                        allow_empty=True,
                    ),
                    "c7_total": travel.total_value,
                    "status_reserva": travel.booking_status,
                    "criticidade": travel.criticality,
                    "status_politica": travel.policy_status,
                    "prioridade": travel.priority,
                }
            )

        return records

    def _write_artifacts(
        self,
        records: list[dict[str, str | float]],
    ) -> ERPExportResult:
        json_file = self.output_dir / f"{ERP_BASENAME}.json"
        csv_file = self.output_dir / f"{ERP_BASENAME}.csv"
        checksum_file = self.output_dir / f"{ERP_BASENAME}.json.sha256"
        targets = (json_file, csv_file, checksum_file)
        existing = [path.name for path in targets if path.exists()]
        if existing:
            raise AutomationError(
                "A exportação ERP não sobrescreve artefatos existentes: "
                + ", ".join(existing)
            )
        if self.output_dir.exists() and not self.output_dir.is_dir():
            raise AutomationError(
                f"O destino da exportação ERP não é uma pasta: {self.output_dir}"
            )

        self.output_dir.mkdir(parents=True, exist_ok=True)
        with TemporaryDirectory(
            dir=self.output_dir,
            prefix=".erp-export-",
        ) as temporary:
            temporary_dir = Path(temporary)
            temporary_json = temporary_dir / json_file.name
            temporary_csv = temporary_dir / csv_file.name
            temporary_checksum = temporary_dir / checksum_file.name

            with temporary_json.open("w", encoding="utf-8", newline="\n") as file:
                json.dump(records, file, ensure_ascii=False, indent=2)
                file.write("\n")

            with temporary_csv.open(
                "w",
                encoding="utf-8-sig",
                newline="",
            ) as file:
                writer = csv.DictWriter(
                    file,
                    fieldnames=ERP_FIELDS,
                    delimiter=";",
                    lineterminator="\n",
                )
                writer.writeheader()
                writer.writerows(records)

            digest = _sha256(temporary_json)
            temporary_checksum.write_text(
                f"{digest}  {json_file.name}\n",
                encoding="utf-8",
            )

            temporary_json.replace(json_file)
            temporary_csv.replace(csv_file)
            temporary_checksum.replace(checksum_file)

        logger.info(
            "Exportação ERP concluída: %d registros; SHA-256=%s",
            len(records),
            digest,
        )
        return ERPExportResult(
            json_file=json_file,
            csv_file=csv_file,
            checksum_file=checksum_file,
            record_count=len(records),
            sha256=digest,
        )


def _iso_date(value: date | datetime) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return value.isoformat()


def _cell_text(
    worksheet: object,
    columns: dict[str, int],
    row: int,
    field: str,
) -> str:
    column = columns.get(field)
    if column is None:
        return ""
    value = worksheet.cell(row, column).value
    return "" if value is None else str(value).strip()


def _cell_number(
    worksheet: object,
    columns: dict[str, int],
    row: int,
    field: str,
    description: str,
    *,
    allow_empty: bool = False,
) -> float:
    value = worksheet.cell(row, columns[field]).value
    if value in (None, "") and allow_empty:
        return 0.0
    if isinstance(value, bool) or not isinstance(value, Real):
        raise ValidationError(
            f"Linha {row}: o campo '{description}' deve ser numérico. Valor: {value!r}"
        )
    return float(value)


def _sha256(file_path: Path) -> str:
    digest = hashlib.sha256()
    with file_path.open("rb") as file:
        for block in iter(lambda: file.read(65536), b""):
            digest.update(block)
    return digest.hexdigest()

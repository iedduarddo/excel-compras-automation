"""Planejamento e execução segura de transformações genéricas em planilhas."""

from __future__ import annotations

import json
import re
import shutil
from collections import Counter
from dataclasses import asdict, dataclass, replace
from datetime import UTC, date, datetime
from enum import StrEnum
from hashlib import sha256
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.assistant.workspace import AssistantWorkspace
from src.core.exceptions import AutomationError
from src.services.text import normalize_text, sanitize_filename, text_similarity

_PLAN_ID = re.compile(r"^[0-9a-f]{12}$")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_MUTATING_ACTIONS = {"limpar", "organizar", "calcular", "resumir", "relatorio"}


class UniversalAction(StrEnum):
    """Ações permitidas pelo executor universal."""

    CLEAN = "limpar"
    ORGANIZE = "organizar"
    CALCULATE = "calcular"
    SUMMARIZE = "resumir"
    REPORT = "relatorio"
    GENERATE_ADAPTER = "gerar_adaptador"


@dataclass(frozen=True, slots=True)
class ColumnProfile:
    """Descrição observada de uma coluna tabular."""

    header: str
    column: int
    kind: str
    non_empty: int
    unique_values: int


@dataclass(frozen=True, slots=True)
class TableProfile:
    """Área tabular inferida sem depender de um modelo de negócio."""

    sheet: str
    header_row: int
    data_start_row: int
    data_end_row: int
    data_rows: int
    columns: tuple[ColumnProfile, ...]
    blank_cells: int
    duplicate_rows: int


@dataclass(frozen=True, slots=True)
class WorkbookProfile:
    """Perfil genérico de uma pasta de trabalho."""

    source: Path
    sheets: tuple[str, ...]
    tables: tuple[TableProfile, ...]


@dataclass(frozen=True, slots=True)
class UniversalPlan:
    """Plano persistido que precisa de confirmação antes da execução."""

    plan_id: str
    source: Path
    source_sha256: str
    created_at: str
    status: str
    actions: tuple[UniversalAction, ...]
    options: dict[str, Any]
    profile: WorkbookProfile
    preview: tuple[str, ...]
    warnings: tuple[str, ...]
    output_file: Path | None
    backup_file: Path | None
    adapter_file: Path | None


@dataclass(frozen=True, slots=True)
class UniversalExecutionResult:
    """Artefatos produzidos por um plano confirmado."""

    plan_id: str
    source: Path
    output_file: Path | None = None
    backup_file: Path | None = None
    adapter_file: Path | None = None


class UniversalAutomation:
    """Cria prévias auditáveis e aplica somente planos confirmados."""

    def __init__(self, workspace: AssistantWorkspace) -> None:
        self.workspace = workspace

    def profile(self, source: Path) -> WorkbookProfile:
        """Reconhece tabelas e tipos em uma estrutura ainda desconhecida."""

        self.workspace.ensure()
        source = self._resolve_source(source)
        workbook = load_workbook(
            source,
            data_only=False,
            read_only=True,
            keep_vba=source.suffix.casefold() == ".xlsm",
        )
        try:
            tables = tuple(
                profile
                for worksheet in workbook.worksheets
                if (profile := _profile_worksheet(worksheet)) is not None
            )
            if not tables:
                raise AutomationError(
                    "Nenhuma tabela com cabeçalho e dados foi reconhecida. "
                    "Revise a planilha ou informe uma estrutura mais regular."
                )
            return WorkbookProfile(
                source=source,
                sheets=tuple(worksheet.title for worksheet in workbook.worksheets),
                tables=tables,
            )
        finally:
            workbook.close()

    def create_plan(
        self,
        source: Path,
        actions: tuple[UniversalAction, ...],
        *,
        options: dict[str, Any] | None = None,
    ) -> tuple[UniversalPlan, Path, Path]:
        """Persiste plano e prévia sem modificar a planilha de entrada."""

        if not actions:
            raise AutomationError("O pedido não contém uma ação universal permitida.")
        self.workspace.ensure()
        profile = self.profile(source)
        options = dict(options or {})
        self._resolve_requested_column(profile, options)
        if (
            UniversalAction.CALCULATE in actions
            and options.get("resolved_kind")
            and options["resolved_kind"] not in {"inteiro", "decimal"}
        ):
            raise AutomationError(
                f"A coluna '{options['resolved_header']}' não é numérica. "
                "Escolha uma coluna de números para calcular."
            )
        plan_id = _new_plan_id(source, actions)
        source_hash = _file_sha256(source)
        mutates_workbook = any(action.value in _MUTATING_ACTIONS for action in actions)
        output_file = (
            self.workspace.output_dir
            / f"{sanitize_filename(source.stem)}_automatizado_{plan_id}{source.suffix}"
            if mutates_workbook
            else None
        )
        backup_file = (
            self.workspace.backup_dir
            / f"{sanitize_filename(source.stem)}_backup_universal_{plan_id}{source.suffix}"
            if mutates_workbook
            else None
        )
        adapter_file = (
            self.workspace.universal_adapters_dir
            / f"{sanitize_filename(source.stem)}_{plan_id}.json"
            if UniversalAction.GENERATE_ADAPTER in actions
            else None
        )
        preview, warnings = _build_preview(profile, actions, options)
        plan = UniversalPlan(
            plan_id=plan_id,
            source=source,
            source_sha256=source_hash,
            created_at=datetime.now(UTC).isoformat(),
            status="aguardando_confirmacao",
            actions=actions,
            options=options,
            profile=profile,
            preview=preview,
            warnings=warnings,
            output_file=output_file,
            backup_file=backup_file,
            adapter_file=adapter_file,
        )
        json_path = self.workspace.plans_dir / f"{plan_id}.json"
        markdown_path = self.workspace.plans_dir / f"{plan_id}.md"
        json_path.write_text(
            json.dumps(_serialize_plan(plan), ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        markdown_path.write_text(_format_plan_markdown(plan), encoding="utf-8")
        return plan, json_path, markdown_path

    def apply_plan(self, plan_id: str) -> UniversalExecutionResult:
        """Aplica um plano confirmado somente se a origem continuar idêntica."""

        plan = self.load_plan(plan_id)
        if plan.status != "aguardando_confirmacao":
            raise AutomationError(
                f"O plano {plan.plan_id} não aguarda confirmação: {plan.status}."
            )
        if _file_sha256(plan.source) != plan.source_sha256:
            raise AutomationError(
                "A planilha mudou depois da prévia. Gere um novo plano antes de "
                "confirmar."
            )
        for artifact in (plan.output_file, plan.backup_file, plan.adapter_file):
            if artifact is not None and artifact.exists():
                raise AutomationError(
                    f"O artefato previsto já existe e não será sobrescrito: {artifact}"
                )

        mutates_workbook = any(
            action.value in _MUTATING_ACTIONS for action in plan.actions
        )
        try:
            if mutates_workbook:
                if plan.output_file is None or plan.backup_file is None:
                    raise AutomationError(
                        "O plano não contém caminhos seguros de saída."
                    )
                shutil.copy2(plan.source, plan.backup_file)
                shutil.copy2(plan.source, plan.output_file)
                self._apply_workbook_actions(plan)

            if UniversalAction.GENERATE_ADAPTER in plan.actions:
                if plan.adapter_file is None:
                    raise AutomationError("O plano não contém o caminho do adaptador.")
                plan.adapter_file.parent.mkdir(parents=True, exist_ok=True)
                plan.adapter_file.write_text(
                    json.dumps(_adapter_payload(plan), ensure_ascii=False, indent=2)
                    + "\n",
                    encoding="utf-8",
                )
        except Exception:
            for artifact in (plan.output_file, plan.backup_file, plan.adapter_file):
                if artifact is not None:
                    artifact.unlink(missing_ok=True)
            raise

        self._update_plan_status(plan, "concluido")
        return UniversalExecutionResult(
            plan_id=plan.plan_id,
            source=plan.source,
            output_file=plan.output_file,
            backup_file=plan.backup_file,
            adapter_file=plan.adapter_file,
        )

    def cancel_plan(self, plan_id: str) -> UniversalPlan:
        """Cancela um plano ainda não executado."""

        plan = self.load_plan(plan_id)
        if plan.status != "aguardando_confirmacao":
            raise AutomationError(
                f"O plano {plan.plan_id} não pode ser cancelado: {plan.status}."
            )
        self._update_plan_status(plan, "cancelado")
        return self.load_plan(plan.plan_id)

    def load_plan(self, plan_id: str) -> UniversalPlan:
        """Carrega um plano pelo identificador, sem aceitar caminhos livres."""

        normalized = plan_id.strip().casefold()
        if not _PLAN_ID.fullmatch(normalized):
            raise AutomationError("Identificador de plano inválido.")
        plan_path = self.workspace.plans_dir / f"{normalized}.json"
        if not plan_path.is_file():
            raise AutomationError(f"Plano não encontrado: {normalized}")
        try:
            payload = json.loads(plan_path.read_text(encoding="utf-8"))
            plan = _deserialize_plan(payload)
        except (KeyError, TypeError, ValueError, json.JSONDecodeError) as error:
            raise AutomationError(f"O plano {normalized} está corrompido.") from error
        self._validate_plan_paths(plan)
        return plan

    def _resolve_source(self, source: Path) -> Path:
        resolved = source.resolve()
        root = self.workspace.input_dir.resolve()
        if not resolved.is_file() or not resolved.is_relative_to(root):
            raise AutomationError(
                "A automação universal aceita somente arquivos da pasta entrada."
            )
        return resolved

    def _validate_plan_paths(self, plan: UniversalPlan) -> None:
        expected_roots = (
            (plan.source, self.workspace.input_dir),
            (plan.output_file, self.workspace.output_dir),
            (plan.backup_file, self.workspace.backup_dir),
            (plan.adapter_file, self.workspace.universal_adapters_dir),
        )
        for path, root in expected_roots:
            if path is not None and not path.resolve().is_relative_to(root.resolve()):
                raise AutomationError(
                    f"O plano {plan.plan_id} contém um caminho fora da pasta permitida."
                )

    def _resolve_requested_column(
        self,
        profile: WorkbookProfile,
        options: dict[str, Any],
    ) -> None:
        requested = options.get("column") or options.get("sort_column")
        if not requested:
            return
        sheet_name = options.get("sheet")
        matches: list[tuple[float, TableProfile, ColumnProfile]] = []
        for table in profile.tables:
            if sheet_name and normalize_text(table.sheet) != normalize_text(sheet_name):
                continue
            for column in table.columns:
                score = text_similarity(requested, column.header)
                if score >= 0.76:
                    matches.append((score, table, column))
        matches.sort(key=lambda item: item[0], reverse=True)
        if not matches or (len(matches) > 1 and matches[0][0] == matches[1][0]):
            raise AutomationError(
                f"A coluna '{requested}' não foi encontrada de forma única. "
                'Informe também aba="NOME".'
            )
        _, table, column = matches[0]
        options["resolved_sheet"] = table.sheet
        options["resolved_column"] = column.column
        options["resolved_header"] = column.header
        options["resolved_kind"] = column.kind

    def _apply_workbook_actions(self, plan: UniversalPlan) -> None:
        if plan.output_file is None:
            return
        workbook = load_workbook(
            plan.output_file,
            data_only=False,
            keep_vba=plan.output_file.suffix.casefold() == ".xlsm",
        )
        try:
            if UniversalAction.CLEAN in plan.actions:
                _apply_clean(workbook, plan.profile, plan.options)
            if UniversalAction.ORGANIZE in plan.actions:
                _apply_organize(workbook, plan.profile, plan.options)
            if UniversalAction.CALCULATE in plan.actions:
                _apply_calculations(workbook, plan.profile, plan.options)
            if UniversalAction.SUMMARIZE in plan.actions:
                _apply_summary(workbook, plan.profile, plan)
            if UniversalAction.REPORT in plan.actions:
                _apply_report(workbook, plan.profile, plan)
            workbook.calculation.calcMode = "auto"
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
            workbook.save(plan.output_file)
        finally:
            workbook.close()

    def _update_plan_status(self, plan: UniversalPlan, status: str) -> None:
        payload = _serialize_plan(plan)
        payload["status"] = status
        payload["updated_at"] = datetime.now(UTC).isoformat()
        plan_path = self.workspace.plans_dir / f"{plan.plan_id}.json"
        plan_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        markdown_path = self.workspace.plans_dir / f"{plan.plan_id}.md"
        markdown_path.write_text(
            _format_plan_markdown(replace(plan, status=status)),
            encoding="utf-8",
        )


def _profile_worksheet(worksheet: Any) -> TableProfile | None:
    if worksheet.max_row < 2 or worksheet.max_column < 2:
        return None
    header_row = _find_header_row(worksheet)
    if header_row is None:
        return None
    header_cells = [
        (column, worksheet.cell(header_row, column).value)
        for column in range(1, min(worksheet.max_column, 200) + 1)
        if worksheet.cell(header_row, column).value not in (None, "")
    ]
    if len(header_cells) < 2:
        return None
    last_row = _last_data_row(worksheet, header_row, tuple(c for c, _ in header_cells))
    if last_row <= header_row:
        return None

    columns: list[ColumnProfile] = []
    for column, header in header_cells:
        values = [
            worksheet.cell(row, column).value
            for row in range(header_row + 1, last_row + 1)
            if worksheet.cell(row, column).value not in (None, "")
        ]
        columns.append(
            ColumnProfile(
                header=" ".join(str(header).split()),
                column=column,
                kind=_infer_kind(values),
                non_empty=len(values),
                unique_values=len({_stable_value(value) for value in values}),
            )
        )

    row_values = [
        tuple(worksheet.cell(row, column).value for column, _ in header_cells)
        for row in range(header_row + 1, last_row + 1)
    ]
    populated = [
        row for row in row_values if any(value not in (None, "") for value in row)
    ]
    blank_cells = sum(value in (None, "") for row in populated for value in row)
    duplicate_rows = len(populated) - len({_stable_row(row) for row in populated})
    return TableProfile(
        sheet=worksheet.title,
        header_row=header_row,
        data_start_row=header_row + 1,
        data_end_row=last_row,
        data_rows=len(populated),
        columns=tuple(columns),
        blank_cells=blank_cells,
        duplicate_rows=duplicate_rows,
    )


def _find_header_row(worksheet: Any) -> int | None:
    best: tuple[float, int] | None = None
    for row in range(1, min(worksheet.max_row, 30) + 1):
        values = [
            worksheet.cell(row, column).value
            for column in range(1, min(worksheet.max_column, 200) + 1)
        ]
        populated = [value for value in values if value not in (None, "")]
        if len(populated) < 2:
            continue
        strings = sum(isinstance(value, str) for value in populated)
        unique = len({normalize_text(value) for value in populated})
        below = 0
        if row < worksheet.max_row:
            below = sum(
                worksheet.cell(row + 1, column).value not in (None, "")
                for column in range(1, len(values) + 1)
            )
        score = len(populated) * 2 + strings + unique + min(below, len(populated))
        candidate = (float(score), -row)
        if best is None or candidate > (best[0], -best[1]):
            best = (float(score), row)
    return best[1] if best is not None else None


def _last_data_row(worksheet: Any, header_row: int, columns: tuple[int, ...]) -> int:
    for row in range(worksheet.max_row, header_row, -1):
        if any(
            worksheet.cell(row, column).value not in (None, "") for column in columns
        ):
            return row
    return header_row


def _infer_kind(values: list[Any]) -> str:
    kinds = {_value_kind(value) for value in values[:500]}
    if not kinds:
        return "vazio"
    if kinds <= {"inteiro", "decimal"}:
        return "decimal" if "decimal" in kinds else "inteiro"
    return next(iter(kinds)) if len(kinds) == 1 else "misto"


def _value_kind(value: Any) -> str:
    if isinstance(value, str) and value.startswith("="):
        return "formula"
    if isinstance(value, bool):
        return "booleano"
    if isinstance(value, (datetime, date)):
        return "data"
    if isinstance(value, int):
        return "inteiro"
    if isinstance(value, float):
        return "decimal"
    return "texto"


def _stable_value(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def _stable_row(row: tuple[Any, ...]) -> tuple[str, ...]:
    return tuple(_stable_value(value) for value in row)


def _build_preview(
    profile: WorkbookProfile,
    actions: tuple[UniversalAction, ...],
    options: dict[str, Any],
) -> tuple[tuple[str, ...], tuple[str, ...]]:
    rows = sum(table.data_rows for table in profile.tables)
    columns = sum(len(table.columns) for table in profile.tables)
    previews = [
        f"Estrutura: {len(profile.tables)} tabela(s), {rows} linha(s) e "
        f"{columns} coluna(s)."
    ]
    warnings = [
        "O arquivo original será preservado; o resultado será gravado na pasta saida."
    ]
    for action in actions:
        if action is UniversalAction.CLEAN:
            duplicates = sum(table.duplicate_rows for table in profile.tables)
            previews.append(
                "Limpar: normalizar espaços e células textuais vazias"
                + (
                    f"; remover {duplicates} linha(s) duplicada(s)."
                    if options.get("remove_duplicates")
                    else "."
                )
            )
        elif action is UniversalAction.ORGANIZE:
            text = "Organizar: aplicar filtros, congelar cabeçalhos e ajustar larguras."
            if options.get("resolved_header"):
                text += f" Ordenação preparada por '{options['resolved_header']}'."
            previews.append(text)
        elif action is UniversalAction.CALCULATE:
            previews.append(
                "Calcular: criar uma aba com contagem, soma, média, mínimo e máximo "
                "das colunas numéricas selecionadas."
            )
        elif action is UniversalAction.SUMMARIZE:
            previews.append(
                "Resumir: criar uma aba com linhas, colunas, vazios e duplicidades."
            )
        elif action is UniversalAction.REPORT:
            previews.append(
                "Relatório: criar uma aba executiva com indicadores das tabelas."
            )
        elif action is UniversalAction.GENERATE_ADAPTER:
            previews.append(
                "Adaptador: gerar um perfil universal JSON com abas, cabeçalhos e tipos."
            )
    if any(table.duplicate_rows for table in profile.tables) and not options.get(
        "remove_duplicates"
    ):
        warnings.append(
            "Foram encontradas linhas duplicadas; elas só serão removidas quando o "
            "pedido disser 'remover duplicados'."
        )
    return tuple(previews), tuple(warnings)


def _apply_clean(
    workbook: Any, profile: WorkbookProfile, options: dict[str, Any]
) -> None:
    for table in profile.tables:
        worksheet = workbook[table.sheet]
        for column in table.columns:
            header = worksheet.cell(table.header_row, column.column)
            if isinstance(header.value, str):
                header.value = " ".join(header.value.split())
            for row in range(table.data_start_row, table.data_end_row + 1):
                cell = worksheet.cell(row, column.column)
                if isinstance(cell.value, str) and not cell.value.startswith("="):
                    cleaned = " ".join(cell.value.split())
                    cell.value = cleaned or None
        if options.get("remove_duplicates"):
            seen: set[tuple[str, ...]] = set()
            duplicates: list[int] = []
            for row in range(table.data_start_row, table.data_end_row + 1):
                values = tuple(
                    worksheet.cell(row, column.column).value for column in table.columns
                )
                if not any(value not in (None, "") for value in values):
                    continue
                key = _stable_row(values)
                if key in seen:
                    duplicates.append(row)
                else:
                    seen.add(key)
            for row in reversed(duplicates):
                worksheet.delete_rows(row)


def _apply_organize(
    workbook: Any,
    profile: WorkbookProfile,
    options: dict[str, Any],
) -> None:
    for table in profile.tables:
        worksheet = workbook[table.sheet]
        first_column = min(column.column for column in table.columns)
        last_column = max(column.column for column in table.columns)
        end_row = worksheet.max_row
        worksheet.freeze_panes = worksheet.cell(table.data_start_row, first_column)
        worksheet.auto_filter.ref = (
            f"{get_column_letter(first_column)}{table.header_row}:"
            f"{get_column_letter(last_column)}{end_row}"
        )
        for column in table.columns:
            cell = worksheet.cell(table.header_row, column.column)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            samples = [
                str(worksheet.cell(row, column.column).value or "")
                for row in range(
                    table.header_row,
                    min(worksheet.max_row, table.header_row + 200) + 1,
                )
            ]
            width = min(max(max(map(len, samples), default=0) + 2, 10), 45)
            worksheet.column_dimensions[get_column_letter(column.column)].width = width

    sheet = options.get("resolved_sheet")
    column = options.get("resolved_column")
    if sheet and column:
        table = next(table for table in profile.tables if table.sheet == sheet)
        worksheet = workbook[sheet]
        ref = (
            f"{get_column_letter(int(column))}{table.data_start_row}:"
            f"{get_column_letter(int(column))}{worksheet.max_row}"
        )
        worksheet.auto_filter.add_sort_condition(
            ref,
            descending=bool(options.get("descending")),
        )


def _apply_calculations(
    workbook: Any,
    profile: WorkbookProfile,
    options: dict[str, Any],
) -> None:
    worksheet = workbook.create_sheet(
        _unique_sheet_title(workbook, "Calculos_Automacao")
    )
    headers = ("Aba", "Coluna", "Registros", "Soma", "Media", "Minimo", "Maximo")
    worksheet.append(headers)
    requested_sheet = options.get("resolved_sheet")
    requested_column = options.get("resolved_column")
    for table in profile.tables:
        if requested_sheet and table.sheet != requested_sheet:
            continue
        source = workbook[table.sheet]
        for column in table.columns:
            if requested_column and column.column != requested_column:
                continue
            values = [
                source.cell(row, column.column).value
                for row in range(table.data_start_row, source.max_row + 1)
            ]
            numbers = [
                float(value)
                for value in values
                if isinstance(value, (int, float)) and not isinstance(value, bool)
            ]
            if not numbers:
                continue
            worksheet.append(
                (
                    table.sheet,
                    column.header,
                    len(numbers),
                    sum(numbers),
                    mean(numbers),
                    min(numbers),
                    max(numbers),
                )
            )
    _style_generated_sheet(worksheet)


def _apply_summary(
    workbook: Any, profile: WorkbookProfile, plan: UniversalPlan
) -> None:
    worksheet = workbook.create_sheet(_unique_sheet_title(workbook, "Resumo_Automacao"))
    worksheet.append(("Resumo da automacao", plan.plan_id))
    worksheet.append(())
    worksheet.append(("Aba", "Linhas", "Colunas", "Celulas vazias", "Duplicidades"))
    for table in profile.tables:
        worksheet.append(
            (
                table.sheet,
                table.data_rows,
                len(table.columns),
                table.blank_cells,
                table.duplicate_rows,
            )
        )
    _style_generated_sheet(worksheet, header_row=3)


def _apply_report(workbook: Any, profile: WorkbookProfile, plan: UniversalPlan) -> None:
    worksheet = workbook.create_sheet(
        _unique_sheet_title(workbook, "Relatorio_Automacao")
    )
    worksheet.append(("RELATORIO DA AUTOMACAO UNIVERSAL",))
    worksheet.append(("Arquivo", plan.source.name))
    worksheet.append(("Plano", plan.plan_id))
    worksheet.append(("Gerado em UTC", datetime.now(UTC).isoformat()))
    worksheet.append(())
    worksheet.append(("Aba", "Registros", "Campos", "Completude", "Tipos observados"))
    for table in profile.tables:
        total = table.data_rows * len(table.columns)
        completeness = 1 - (table.blank_cells / total if total else 0)
        kinds = Counter(column.kind for column in table.columns)
        types = ", ".join(f"{kind}: {count}" for kind, count in sorted(kinds.items()))
        worksheet.append(
            (table.sheet, table.data_rows, len(table.columns), completeness, types)
        )
        worksheet.cell(worksheet.max_row, 4).number_format = "0.0%"
    _style_generated_sheet(worksheet, header_row=6)


def _style_generated_sheet(worksheet: Any, header_row: int = 1) -> None:
    for cell in worksheet[header_row]:
        if cell.value is not None:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
    worksheet.freeze_panes = worksheet.cell(header_row + 1, 1)
    for column in range(1, worksheet.max_column + 1):
        values = [
            str(worksheet.cell(row, column).value or "")
            for row in range(1, worksheet.max_row + 1)
        ]
        worksheet.column_dimensions[get_column_letter(column)].width = min(
            max(max(map(len, values), default=0) + 2, 12),
            50,
        )


def _unique_sheet_title(workbook: Any, base: str) -> str:
    if base not in workbook.sheetnames:
        return base
    counter = 2
    while f"{base}_{counter}" in workbook.sheetnames:
        counter += 1
    return f"{base}_{counter}"


def _new_plan_id(source: Path, actions: tuple[UniversalAction, ...]) -> str:
    payload = f"{source.resolve()}|{datetime.now(UTC).isoformat()}|" + ",".join(
        action.value for action in actions
    )
    return sha256(payload.encode("utf-8")).hexdigest()[:12]


def _file_sha256(path: Path) -> str:
    digest = sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _serialize_plan(plan: UniversalPlan) -> dict[str, Any]:
    profile = asdict(plan.profile)
    profile["source"] = str(plan.profile.source)
    return {
        "schema_version": 1,
        "plan_id": plan.plan_id,
        "source": str(plan.source),
        "source_sha256": plan.source_sha256,
        "created_at": plan.created_at,
        "status": plan.status,
        "actions": [action.value for action in plan.actions],
        "options": plan.options,
        "profile": profile,
        "preview": list(plan.preview),
        "warnings": list(plan.warnings),
        "output_file": str(plan.output_file) if plan.output_file else None,
        "backup_file": str(plan.backup_file) if plan.backup_file else None,
        "adapter_file": str(plan.adapter_file) if plan.adapter_file else None,
    }


def _deserialize_plan(payload: dict[str, Any]) -> UniversalPlan:
    profile_payload = payload["profile"]
    tables = tuple(
        TableProfile(
            sheet=table["sheet"],
            header_row=int(table["header_row"]),
            data_start_row=int(table["data_start_row"]),
            data_end_row=int(table["data_end_row"]),
            data_rows=int(table["data_rows"]),
            columns=tuple(ColumnProfile(**column) for column in table["columns"]),
            blank_cells=int(table["blank_cells"]),
            duplicate_rows=int(table["duplicate_rows"]),
        )
        for table in profile_payload["tables"]
    )
    profile = WorkbookProfile(
        source=Path(profile_payload["source"]).resolve(),
        sheets=tuple(profile_payload["sheets"]),
        tables=tables,
    )
    return UniversalPlan(
        plan_id=payload["plan_id"],
        source=Path(payload["source"]).resolve(),
        source_sha256=payload["source_sha256"],
        created_at=payload["created_at"],
        status=payload["status"],
        actions=tuple(UniversalAction(action) for action in payload["actions"]),
        options=dict(payload["options"]),
        profile=profile,
        preview=tuple(payload["preview"]),
        warnings=tuple(payload["warnings"]),
        output_file=Path(payload["output_file"]) if payload["output_file"] else None,
        backup_file=Path(payload["backup_file"]) if payload["backup_file"] else None,
        adapter_file=Path(payload["adapter_file"]) if payload["adapter_file"] else None,
    )


def _format_plan_markdown(plan: UniversalPlan) -> str:
    if plan.status != "aguardando_confirmacao":
        return "\n".join(
            (
                f"# Plano {plan.plan_id}",
                "",
                f"- Arquivo: `{plan.source.name}`",
                f"- Estado final: `{plan.status}`",
                "",
            )
        )
    lines = [
        f"# Prévia do plano {plan.plan_id}",
        "",
        f"- Arquivo: `{plan.source.name}`",
        f"- SHA-256: `{plan.source_sha256}`",
        f"- Estado: `{plan.status}`",
        f"- Ações: {', '.join(action.value for action in plan.actions)}",
        "",
        "## Alterações previstas",
        "",
        *(f"- {item}" for item in plan.preview),
        "",
        "## Avisos",
        "",
        *(f"- {warning}" for warning in plan.warnings),
        "",
        "## Confirmação",
        "",
        "Nada foi alterado. Para executar exatamente este plano:",
        "",
        f'`confirmar plano="{plan.plan_id}"`',
        "",
        "Para cancelar:",
        "",
        f'`cancelar plano="{plan.plan_id}"`',
        "",
    ]
    return "\n".join(lines)


def _adapter_payload(plan: UniversalPlan) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "kind": "universal_workbook_adapter",
        "name": sanitize_filename(plan.source.stem).casefold(),
        "source_sha256": plan.source_sha256,
        "sheets": {
            table.sheet: {
                "header_row": table.header_row,
                "columns": {
                    normalize_text(column.header).replace(" ", "_"): {
                        "source_header": column.header,
                        "column": column.column,
                        "type": column.kind,
                    }
                    for column in table.columns
                },
            }
            for table in plan.profile.tables
        },
    }

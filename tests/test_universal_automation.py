"""Testes do planejamento e da execução universal com confirmação."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from src.assistant.service import FolderAssistant
from src.assistant.universal import UniversalAction, UniversalAutomation
from src.assistant.workspace import AssistantWorkspace
from src.core.exceptions import AutomationError


def make_unknown_workbook(workspace: AssistantWorkspace) -> Path:
    workspace.ensure()
    path = workspace.input_dir / "clientes.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Cadastro Geral"
    worksheet.append(("  Cliente  ", "Cidade", "Valor", "Data"))
    worksheet.append((" Ana   Lima ", "Recife", 120.5, "2026-08-01"))
    worksheet.append(("Bruno", "São Paulo", 80, "2026-08-02"))
    worksheet.append(("Bruno", "São Paulo", 80, "2026-08-02"))
    workbook.save(path)
    workbook.close()
    return path


def test_profiles_unknown_tabular_structure_without_business_aliases(tmp_path) -> None:
    workspace = AssistantWorkspace(tmp_path / "assistente")
    source = make_unknown_workbook(workspace)

    profile = UniversalAutomation(workspace).profile(source)

    assert profile.sheets == ("Cadastro Geral",)
    assert len(profile.tables) == 1
    table = profile.tables[0]
    assert table.header_row == 1
    assert table.data_rows == 3
    assert table.duplicate_rows == 1
    assert [(column.header, column.kind) for column in table.columns] == [
        ("Cliente", "texto"),
        ("Cidade", "texto"),
        ("Valor", "decimal"),
        ("Data", "texto"),
    ]


def test_plan_creates_preview_but_does_not_change_or_copy_workbook(tmp_path) -> None:
    workspace = AssistantWorkspace(tmp_path / "assistente")
    source = make_unknown_workbook(workspace)
    original = source.read_bytes()

    result = FolderAssistant(workspace).execute(
        'limpar e organizar arquivo="clientes.xlsx" remover duplicados'
    )

    assert result.succeeded is True
    item = result.items[0]
    assert item.status == "confirmacao"
    assert item.plan_id is not None
    assert item.preview_file is not None and item.preview_file.is_file()
    assert "Nada foi alterado" in item.preview_file.read_text(encoding="utf-8")
    assert source.read_bytes() == original
    assert not tuple(workspace.output_dir.iterdir())
    assert not tuple(workspace.backup_dir.iterdir())


def test_confirmed_plan_changes_only_copy_and_preserves_original(tmp_path) -> None:
    workspace = AssistantWorkspace(tmp_path / "assistente")
    source = make_unknown_workbook(workspace)
    original = source.read_bytes()
    assistant = FolderAssistant(workspace)
    preview = assistant.execute(
        'limpe e organize arquivo="clientes.xlsx" remover duplicados'
    )
    plan_id = preview.items[0].plan_id

    result = assistant.execute(f'confirmar plano="{plan_id}"')

    item = result.items[0]
    assert result.succeeded is True
    assert item.output_file is not None and item.output_file.is_file()
    assert source.read_bytes() == original
    assert len(tuple(workspace.backup_dir.glob("*.xlsx"))) == 1
    workbook = load_workbook(item.output_file)
    try:
        worksheet = workbook["Cadastro Geral"]
        assert worksheet["A1"].value == "Cliente"
        assert worksheet["A2"].value == "Ana Lima"
        assert worksheet.max_row == 3
        assert worksheet.freeze_panes == "A2"
        assert worksheet.auto_filter.ref == "A1:D3"
    finally:
        workbook.close()


def test_calculate_summarize_and_report_create_reviewable_sheets(tmp_path) -> None:
    workspace = AssistantWorkspace(tmp_path / "assistente")
    make_unknown_workbook(workspace)
    assistant = FolderAssistant(workspace)
    preview = assistant.execute(
        'calcule, resuma e crie relatório arquivo="clientes.xlsx" coluna="Valor"'
    )

    result = assistant.execute(f'confirmar plano="{preview.items[0].plan_id}"')

    workbook = load_workbook(result.items[0].output_file, data_only=True)
    try:
        assert "Calculos_Automacao" in workbook.sheetnames
        assert "Resumo_Automacao" in workbook.sheetnames
        assert "Relatorio_Automacao" in workbook.sheetnames
        calculations = workbook["Calculos_Automacao"]
        assert calculations["B2"].value == "Valor"
        assert calculations["C2"].value == 3
        assert calculations["D2"].value == pytest.approx(280.5)
    finally:
        workbook.close()


def test_generate_adapter_is_also_guarded_by_confirmation(tmp_path) -> None:
    workspace = AssistantWorkspace(tmp_path / "assistente")
    make_unknown_workbook(workspace)
    assistant = FolderAssistant(workspace)
    preview = assistant.execute('gere adaptador arquivo="clientes.xlsx"')

    assert not tuple(workspace.universal_adapters_dir.glob("*.json"))

    result = assistant.execute(f'confirmar plano="{preview.items[0].plan_id}"')

    adapter = result.items[0].adapter_file
    assert adapter is not None and adapter.is_file()
    payload = json.loads(adapter.read_text(encoding="utf-8"))
    assert payload["kind"] == "universal_workbook_adapter"
    assert payload["sheets"]["Cadastro Geral"]["columns"]["valor"] == {
        "source_header": "Valor",
        "column": 3,
        "type": "decimal",
    }


def test_confirmation_stops_when_source_changed_after_preview(tmp_path) -> None:
    workspace = AssistantWorkspace(tmp_path / "assistente")
    source = make_unknown_workbook(workspace)
    universal = UniversalAutomation(workspace)
    plan, _, _ = universal.create_plan(source, (UniversalAction.REPORT,))
    source.write_bytes(source.read_bytes() + b"alterado")

    with pytest.raises(AutomationError, match="mudou depois da prévia"):
        universal.apply_plan(plan.plan_id)


def test_cancelled_plan_cannot_be_executed(tmp_path) -> None:
    workspace = AssistantWorkspace(tmp_path / "assistente")
    source = make_unknown_workbook(workspace)
    universal = UniversalAutomation(workspace)
    plan, _, _ = universal.create_plan(source, (UniversalAction.REPORT,))

    cancelled = universal.cancel_plan(plan.plan_id)

    assert cancelled.status == "cancelado"
    preview = workspace.plans_dir / f"{plan.plan_id}.md"
    assert "Estado final: `cancelado`" in preview.read_text(encoding="utf-8")
    with pytest.raises(AutomationError, match="não aguarda confirmação"):
        universal.apply_plan(plan.plan_id)


def test_calculation_rejects_requested_text_column_during_preview(tmp_path) -> None:
    workspace = AssistantWorkspace(tmp_path / "assistente")
    make_unknown_workbook(workspace)

    with pytest.raises(AutomationError, match="não é numérica"):
        FolderAssistant(workspace).execute(
            'calcular arquivo="clientes.xlsx" coluna="Cidade"'
        )

    assert not tuple(workspace.plans_dir.glob("*.json"))


def test_tampered_plan_cannot_write_outside_workspace(tmp_path) -> None:
    workspace = AssistantWorkspace(tmp_path / "assistente")
    source = make_unknown_workbook(workspace)
    universal = UniversalAutomation(workspace)
    plan, plan_file, _ = universal.create_plan(source, (UniversalAction.REPORT,))
    payload = json.loads(plan_file.read_text(encoding="utf-8"))
    payload["output_file"] = str(tmp_path / "fora.xlsx")
    plan_file.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(AutomationError, match="fora da pasta permitida"):
        universal.apply_plan(plan.plan_id)

    assert not (tmp_path / "fora.xlsx").exists()

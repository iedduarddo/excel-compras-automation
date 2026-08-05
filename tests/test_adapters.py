"""Testes dos perfis de mapeamento de planilhas externas."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from src.core.exceptions import AutomationError
from src.excel.detection import WorkbookDetector
from src.settings import load_aliases, merge_aliases


def write_adapter(path: Path, aliases: dict[str, object]) -> Path:
    path.write_text(
        json.dumps({"name": "origem-teste", "aliases": aliases}, ensure_ascii=False),
        encoding="utf-8",
    )
    return path


def test_adapter_extends_default_aliases_without_mutating_them(tmp_path: Path) -> None:
    default_aliases = load_aliases()
    adapter = write_adapter(
        tmp_path / "adapter.json",
        {
            "sheets": {"base": ["Pedidos Corporativos"]},
            "base_columns": {"request_id": ["Número do Pedido", "ID"]},
        },
    )

    adapted = load_aliases(adapter)

    assert "Pedidos Corporativos" in adapted["sheets"]["base"]
    assert "Número do Pedido" in adapted["base_columns"]["request_id"]
    assert adapted["base_columns"]["request_id"].count("ID") == 1
    assert "Pedidos Corporativos" not in default_aliases["sheets"]["base"]


@pytest.mark.parametrize(
    ("profile", "message"),
    [
        ({}, "objeto JSON chamado 'aliases'"),
        ({"aliases": {"grupo_inexistente": {}}}, "Grupo de aliases desconhecido"),
        (
            {"aliases": {"base_columns": {"campo_inexistente": ["Campo"]}}},
            "Campo canônico desconhecido",
        ),
        (
            {"aliases": {"base_columns": {"request_id": []}}},
            "lista não vazia",
        ),
    ],
)
def test_adapter_rejects_invalid_schema(
    profile: dict[str, object],
    message: str,
) -> None:
    with pytest.raises(AutomationError, match=message):
        merge_aliases(load_aliases(), profile)


def test_adapter_enables_detection_of_external_names(tmp_path: Path) -> None:
    adapter = write_adapter(
        tmp_path / "adapter.json",
        {
            "sheets": {
                "base": ["Pedidos da Agência"],
                "policies": ["Parâmetros da Agência"],
                "responses": ["Indicadores da Agência"],
            },
            "base_columns": {"request_id": ["Número do Pedido"]},
        },
    )
    aliases = load_aliases(adapter)
    workbook = Workbook()
    base = workbook.active
    base.title = "Pedidos da Agência"
    base_headers = [
        aliases["base_columns"][field][0]
        for field in (
            "request_id",
            "request_date",
            "travel_date",
            "cost_center",
            "service_type",
            "supplier",
            "quantity",
            "unit_value",
            "fees",
            "booking_status",
            "criticality",
            "card_status",
        )
    ]
    base_headers[0] = "Número do Pedido"
    base.append(base_headers)

    policies = workbook.create_sheet("Parâmetros da Agência")
    policies.append(
        [
            aliases["policy_columns"][field][0]
            for field in (
                "service_type",
                "limit_value",
                "min_lead_days",
            )
        ]
    )
    responses = workbook.create_sheet("Indicadores da Agência")
    responses.append(
        [aliases["response_columns"][field][0] for field in ("indicator", "answer")]
    )

    layout = WorkbookDetector(aliases).detect(workbook)

    assert layout.base.title == "Pedidos da Agência"
    assert layout.base.columns["request_id"] == 1
    assert layout.policies.title == "Parâmetros da Agência"
    assert layout.responses.title == "Indicadores da Agência"

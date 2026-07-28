"""Testes do escritor usando somente workbooks OpenPyXL em memória."""

from __future__ import annotations

from datetime import date
from hashlib import sha256
from inspect import signature
from pathlib import Path
from unittest.mock import Mock

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

import src.excel.workbook_writer as writer
from src.core.models import Policy, SheetLayout, TravelResult, WorkbookLayout
from src.excel.workbook_writer import (
    _absolute_range,
    _cell,
    _find_pivot_title_row,
    _find_priority_table,
    _prepare_pivot_area,
    _quoted_sheet,
    apply_conditional_formatting,
    create_fallback_summary_and_chart,
    create_support_sheet,
    ensure_derived_columns,
    finalize_workbook,
    load_source_workbook,
    write_base_formulas,
    write_responses,
)
from src.settings import load_aliases, load_rules

BASE_COLUMNS = {
    "request_id": 1,
    "request_date": 2,
    "travel_date": 3,
    "cost_center": 4,
    "service_type": 5,
    "supplier": 6,
    "quantity": 7,
    "unit_value": 8,
    "fees": 9,
    "booking_status": 10,
    "criticality": 11,
    "card_status": 12,
}
POLICY_COLUMNS = {
    "service_type": 1,
    "limit_value": 2,
    "min_lead_days": 3,
}


def make_travel(
    request_id: str,
    *,
    row: int,
    cost_center: str,
    service_type: str,
    supplier: str,
    total: float,
    score: float,
    priority: str,
) -> TravelResult:
    return TravelResult(
        worksheet_row=row,
        request_id=request_id,
        request_date=date(2026, 1, 1),
        travel_date=date(2026, 1, 10),
        service_type=service_type,
        supplier=supplier,
        cost_center=cost_center,
        booking_status="Confirmada",
        criticality="Normal",
        card_status="Conferido",
        total_value=total,
        lead_days=9,
        policy_limit=1000.0,
        min_lead_days=5,
        policy_status="OK",
        limit_difference=total - 1000,
        score=score,
        priority=priority,
        reasons=["custo elevado", "prazo curto"],
        recommended_actions=["renegociar tarifa", "obter aprovação"],
    )


def make_writer_fixture() -> tuple[
    Workbook,
    WorkbookLayout,
    list[TravelResult],
    dict[str, Policy],
    dict[str, object],
    dict[str, object],
]:
    workbook = Workbook()
    base = workbook.active
    base.title = "Base Viagens"
    for field, column in BASE_COLUMNS.items():
        base.cell(1, column, field)
    base.append(
        [
            "VIA-001",
            date(2026, 1, 1),
            date(2026, 1, 10),
            "CC 100",
            "Aéreo",
            "Fornecedor Á",
            1,
            1200,
            50,
            "Confirmada",
            "Normal",
            "Conferido",
        ]
    )
    base.append(
        [
            "VIA-002",
            date(2026, 1, 2),
            date(2026, 1, 12),
            "CC 200",
            "Hotel",
            "Fornecedor B",
            2,
            400,
            20,
            "Pendente",
            "Executivo",
            "Pendente",
        ]
    )
    for row in range(2, 4):
        base.cell(row, 12).fill = PatternFill("solid", fgColor="FFABCDEF")

    policies_sheet = workbook.create_sheet("Políticas O'Brien")
    for field, column in POLICY_COLUMNS.items():
        policies_sheet.cell(1, column, field)
    policies_sheet.append(["Aéreo", 1000, 7])
    policies_sheet.append(["Hotel", 900, 5])

    responses = workbook.create_sheet("Respostas")
    aliases = load_aliases()
    for row, options in enumerate(aliases["indicator_labels"].values(), start=2):
        responses.cell(row, 1, options[0])
    priority_row = 15
    for column, header in enumerate(
        [
            "ID Solicitação",
            "Motivo da prioridade",
            "Ação recomendada",
            "Ordem (1 a 5)",
        ],
        start=1,
    ):
        responses.cell(priority_row, column, header)
    responses.cell(25, 1, "Tabela dinâmica e gráfico")

    layout = WorkbookLayout(
        base=SheetLayout("Base Viagens", 1, BASE_COLUMNS),
        policies=SheetLayout("Políticas O'Brien", 1, POLICY_COLUMNS),
        responses=SheetLayout(
            "Respostas",
            1,
            {"indicator": 1, "answer": 2, "formula_used": 3},
        ),
    )
    travels = [
        make_travel(
            "VIA-001",
            row=2,
            cost_center="CC 100",
            service_type="Aéreo",
            supplier="Fornecedor Á",
            total=1250,
            score=90,
            priority="Crítica",
        ),
        make_travel(
            "VIA-002",
            row=3,
            cost_center="CC 200",
            service_type="Hotel",
            supplier="Fornecedor B",
            total=820,
            score=45,
            priority="Alta",
        ),
    ]
    policies = {
        "aereo": Policy("Aéreo", 1000, 7, 2),
        "hotel": Policy("Hotel", 900, 5, 3),
    }
    return workbook, layout, travels, policies, aliases, load_rules()


@pytest.mark.parametrize(
    ("filename", "keep_vba"),
    [("entrada.xlsx", False), ("ENTRADA.XLSM", True)],
)
def test_load_source_workbook_preserves_macros_only_when_needed(
    filename: str,
    keep_vba: bool,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    expected = object()
    mocked_load = Mock(return_value=expected)
    monkeypatch.setattr(writer, "load_workbook", mocked_load)

    result = load_source_workbook(Path(filename))

    assert result is expected
    mocked_load.assert_called_once_with(
        Path(filename),
        data_only=False,
        keep_vba=keep_vba,
    )


def test_complete_writer_pipeline_serializes_auditable_workbook(
    tmp_path: Path,
) -> None:
    workbook, layout, travels, policies, aliases, rules = make_writer_fixture()

    updated_layout = ensure_derived_columns(workbook, layout, last_row=3)
    repeated_layout = ensure_derived_columns(workbook, updated_layout, last_row=3)
    support_refs = create_support_sheet(
        workbook,
        repeated_layout,
        travels,
        rules,
        last_row=3,
    )
    write_base_formulas(
        workbook,
        repeated_layout,
        policies,
        support_refs,
        last_row=3,
    )
    pivot_start_row = write_responses(
        workbook,
        repeated_layout,
        aliases,
        travels,
        support_refs,
        last_row=3,
        top_quantity=5,
    )
    apply_conditional_formatting(workbook, repeated_layout, last_row=3)
    create_fallback_summary_and_chart(
        workbook,
        repeated_layout,
        travels,
        pivot_start_row,
        last_row=3,
    )
    finalize_workbook(workbook)

    output_file = tmp_path / "resultado.xlsx"
    workbook.save(output_file)
    workbook.close()
    reopened = load_workbook(output_file, data_only=False)
    try:
        base = reopened["Base Viagens"]
        responses = reopened["Respostas"]
        support = reopened["Apoio_Automacao"]
        columns = repeated_layout.base.columns

        assert repeated_layout == updated_layout
        assert layout.base.columns == BASE_COLUMNS
        assert len(base.tables) == 1
        assert next(iter(base.tables.values())).ref == "A1:S3"
        assert columns["total_value"] == 13
        assert columns["priority_score"] == 19
        assert base.column_dimensions["S"].hidden is True
        assert base["M2"].value.startswith("=IFERROR(")
        assert "INDEX(" in base["O2"].value
        assert base["S2"].value.startswith("=ROUND(")
        assert base["R1"].comment is not None
        assert len(base.conditional_formatting._cf_rules) == 2
        assert base.freeze_panes == "A2"

        assert support.sheet_state == "hidden"
        assert support["K3"].value == "Centro de Custo"
        assert support["M4"].value == 1250
        assert support_refs["critical_threshold"].startswith("'Apoio_Automacao'!$B$")

        assert responses["B2"].value.startswith("=SUM(")
        assert responses["C2"].value == "SOMA"
        assert responses["A16"].value == "VIA-001"
        assert responses["D16"].value == 1
        assert responses["D16"].comment is not None
        assert responses.cell(pivot_start_row, 1).value == "Centro de Custo"
        assert responses.cell(pivot_start_row, 2).value == "Aéreo"
        assert len(responses._charts) == 1
        assert responses._charts[0].title is not None

        assert reopened.calculation.calcMode == "auto"
        assert reopened.calculation.fullCalcOnLoad is True
        assert reopened.calculation.forceFullCalc is True
    finally:
        reopened.close()


def test_support_sheet_is_replaced_and_sorted_deterministically() -> None:
    workbook, layout, travels, _, _, rules = make_writer_fixture()
    layout = ensure_derived_columns(workbook, layout, last_row=3)
    old_support = workbook.create_sheet("Apoio_Automacao")
    old_support["A1"] = "conteúdo antigo"
    travels.append(
        make_travel(
            "VIA-003",
            row=4,
            cost_center="cc 100",
            service_type="Hotel",
            supplier="fornecedor á",
            total=100,
            score=10,
            priority="Normal",
        )
    )

    references = create_support_sheet(workbook, layout, travels, rules, last_row=3)

    support = workbook["Apoio_Automacao"]
    assert support["A1"].value != "conteúdo antigo"
    cost_centers = [support["E4"].value, support["E5"].value, support["E6"].value]
    assert set(cost_centers) == {"CC 100", "cc 100", "CC 200"}
    assert cost_centers[-1] == "CC 200"
    assert support["F4"].value.startswith("=SUMIF(")
    assert references["cost_center_labels"].endswith("$E$6")
    assert references["supplier_labels"].endswith("$H$6")


def test_writer_search_and_reference_helpers_cover_fallbacks() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "O'Brien"
    worksheet["A1"] = "conteúdo"
    worksheet.merge_cells("A5:C5")

    priority_row, columns = _find_priority_table(worksheet)
    pivot_title_row = _find_pivot_title_row(worksheet)
    _prepare_pivot_area(
        worksheet,
        title_row=3,
        pivot_start_row=5,
    )

    assert priority_row == 7
    assert columns == {"request_id": 1, "reason": 2, "action": 3, "order": 4}
    assert pivot_title_row == 7
    assert "A5:C5" not in {str(item) for item in worksheet.merged_cells.ranges}
    assert _quoted_sheet("O'Brien") == "'O''Brien'"
    assert _absolute_range("O'Brien", 28, 2, 10) == "'O''Brien'!$AB$2:$AB$10"
    assert _cell(28, 10) == "AB10"


def test_writer_public_api_signatures_are_stable() -> None:
    expected = {
        "load_source_workbook": "(path: 'Path') -> 'Workbook'",
        "ensure_derived_columns": (
            "(workbook: 'Workbook', layout: 'WorkbookLayout', "
            "last_row: 'int') -> 'WorkbookLayout'"
        ),
        "create_support_sheet": (
            "(workbook: 'Workbook', layout: 'WorkbookLayout', "
            "travels: 'list[TravelResult]', rules: 'dict[str, object]', "
            "last_row: 'int') -> 'dict[str, str]'"
        ),
        "write_base_formulas": (
            "(workbook: 'Workbook', layout: 'WorkbookLayout', "
            "policies: 'dict[str, Policy]', support_refs: 'dict[str, str]', "
            "last_row: 'int') -> 'None'"
        ),
        "write_responses": (
            "(workbook: 'Workbook', layout: 'WorkbookLayout', "
            "aliases: 'dict[str, object]', travels: 'list[TravelResult]', "
            "support_refs: 'dict[str, str]', last_row: 'int', "
            "top_quantity: 'int') -> 'int'"
        ),
        "create_fallback_summary_and_chart": (
            "(workbook: 'Workbook', layout: 'WorkbookLayout', "
            "travels: 'list[TravelResult]', pivot_start_row: 'int', "
            "last_row: 'int') -> 'None'"
        ),
        "apply_conditional_formatting": (
            "(workbook: 'Workbook', layout: 'WorkbookLayout', "
            "last_row: 'int') -> 'None'"
        ),
        "finalize_workbook": "(workbook: 'Workbook') -> 'None'",
    }

    actual = {name: str(signature(getattr(writer, name))) for name in expected}

    assert actual == expected


def test_writer_formula_and_layout_contract_is_stable() -> None:
    workbook, layout, travels, policies, aliases, rules = make_writer_fixture()
    try:
        layout = ensure_derived_columns(workbook, layout, last_row=3)
        support_refs = create_support_sheet(
            workbook,
            layout,
            travels,
            rules,
            last_row=3,
        )
        write_base_formulas(
            workbook,
            layout,
            policies,
            support_refs,
            last_row=3,
        )
        pivot_start_row = write_responses(
            workbook,
            layout,
            aliases,
            travels,
            support_refs,
            last_row=3,
            top_quantity=5,
        )
        apply_conditional_formatting(workbook, layout, last_row=3)
        create_fallback_summary_and_chart(
            workbook,
            layout,
            travels,
            pivot_start_row,
            last_row=3,
        )
        finalize_workbook(workbook)

        base = workbook[layout.base.title]
        responses = workbook[layout.responses.title]
        formula_cells = [f"{column}{row}" for row in (2, 3) for column in "MNOPQRS"]
        formula_hashes = {
            cell: sha256(base[cell].value.encode("utf-8")).hexdigest()
            for cell in formula_cells
        }
        indicator_hashes = {
            f"B{row}": sha256(responses.cell(row, 2).value.encode("utf-8")).hexdigest()
            for row in range(2, 12)
        }

        assert formula_hashes == {
            "M2": "515c398a6c1931fdb83e767cbadb64b06df351fdf1871b1abf7ffabd9e4feb0b",
            "N2": "2fcd1e2a13adb4e0ba820ce106bc1414f0b14be0ccd2b4d1143ace638bc57b90",
            "O2": "ada120a0e6b2987cebf95756449c84f8ba94cf039aa50445ad86405ba45de933",
            "P2": "c14a5cfb69e5377955ce57d7cb35a430cc247ed5379ebdf206f23beacb1cf2b0",
            "Q2": "29d5fee7de153fa79447287392bf81fde1be2e99d555dbc88604f10a2c9ffc57",
            "R2": "41e7f1c345b68cfb4537b74326a2196f021c4093f9c79dfb5d028519a7e5f84e",
            "S2": "1f84d0b098bb35937b2f4d4c77890fd0a0c095a473037b5949351c5537c749ad",
            "M3": "aac759ebf1406a0fe1c77358aeeb70bae31c98af7badc19533f65513a2b61a25",
            "N3": "a88ba3e8201565c1322b56002407484a2195a8107036ede512101c9211157525",
            "O3": "e0c418cf6efeb4abd00b1ee86fc3332d7b85ea738a169b5d692a6996bc0bfc12",
            "P3": "80cc3b4b1a47c4affb18e51a43207986662ad686d075a804b9fa440660415ee8",
            "Q3": "2b41b5ecd576b324f3000bfe1b271c6c8e5066edddb60bd05712a6670ac903d6",
            "R3": "255c2c8dcbc93a066d267fd5549f39a05b622fc35edb4616ad1ac6eaad9e4223",
            "S3": "8272002eefaf22ba1d334660f0b043355f583b6d48649dc158a7ab8161007735",
        }
        assert indicator_hashes == {
            "B2": "09895058db1b806084f402d972c47c7532c86f352f73f6203acd7464c780fc65",
            "B3": "36ec3a5f14de5f55d7e8a11f46ce68e84846ab03e0eb8372d538a70da594c47c",
            "B4": "4f0274fd762be2a68035dd7e1832c80613aad9ba4d297bf0d5820c0f3f1ba6a9",
            "B5": "df06e1aef8bd182d15b1cbf950557715addc76c11ed946d9a7b1803b36b8ef6e",
            "B6": "8dfc0a8e974e87791a0d100e6f286b01c61b47f297b3a8aae3f4dc8f0256a492",
            "B7": "266123f8440331e491635d22dc7944fd8b9e4a1a56c03bafe569d73e9ebeab25",
            "B8": "fbbcb3d9c3e9368ba70441548bae029e11e04e60855397adb23163cbb157f94b",
            "B9": "97a3a33e06bb6de2f6437a287c0c484905f6be8c2a48f8c17e34f1c8cb55a071",
            "B10": "769134e0293633e3141e7f0b39fa18c56277b73527fbd29c04487a5cce1c3d3b",
            "B11": "4ea8a7a434d9ea4f7c9598047f0febd35959ec73ecad688827df40eb11b70fd4",
        }

        conditional_formats = {
            str(cell_range.sqref): [
                (rule.type, tuple(rule.formula or [])) for rule in format_rules
            ]
            for cell_range, format_rules in base.conditional_formatting._cf_rules.items()
        }
        assert conditional_formats == {
            "A2:R3": [("expression", ('$R2="Crítica"',))],
            "P2:Q3": [("expression", ('$P2="Fora"',))],
        }

        chart = responses._charts[0]
        assert pivot_start_row == 27
        assert [series.val.numRef.f for series in chart.ser] == [
            "'Respostas'!$B$28:$B$29",
            "'Respostas'!$C$28:$C$29",
        ]
        assert [series.cat.numRef.f for series in chart.ser] == [
            "'Respostas'!$A$28:$A$29",
            "'Respostas'!$A$28:$A$29",
        ]
        assert sorted(str(item) for item in responses.merged_cells.ranges) == [
            "A25:H25",
            "A26:F26",
        ]
    finally:
        workbook.close()

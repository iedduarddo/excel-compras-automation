import pytest
from openpyxl import Workbook

from src.core.exceptions import DetectionError
from src.excel.detection import (
    BASE_REQUIRED_FIELDS,
    HeaderDetector,
    WorkbookDetector,
    find_indicator_rows,
)
from src.settings import load_aliases


def test_detects_renamed_sheets_and_reordered_columns() -> None:
    workbook = Workbook()
    base = workbook.active
    base.title = "Dados Operacionais 2026"
    base.append(["Relatório interno"])
    base.append([])
    base.append(
        [
            "Prestador",
            "Data do Serviço",
            "ID da Solicitação",
            "Data da Solicitação",
            "Centro Custo",
            "Tipo Serviço",
            "Qtde",
            "Preço Unitário",
            "Encargos",
            "Status da Reserva",
            "Nível de Criticidade",
            "Situação do Cartão",
        ]
    )

    policies = workbook.create_sheet("Parâmetros Corporativos")
    policies.append(["Título"])
    policies.append([])
    policies.append([])
    policies.append(
        [
            "Mínimo de Dias",
            "Teto",
            "Categoria do Serviço",
        ]
    )

    responses = workbook.create_sheet("Painel Gerencial")
    responses.append(["Indicador", "Resultado", "Cálculo"])

    layout = WorkbookDetector(load_aliases()).detect(workbook)

    assert layout.base.title == "Dados Operacionais 2026"
    assert layout.base.header_row == 3
    assert layout.base.columns["request_id"] == 3
    assert layout.base.columns["supplier"] == 1
    assert layout.policies.title == "Parâmetros Corporativos"
    assert layout.policies.header_row == 4
    assert layout.policies.columns["min_lead_days"] == 1
    assert layout.responses.title == "Painel Gerencial"
    assert layout.responses.columns["answer"] == 2


def test_find_indicator_rows_accepts_moved_aliases() -> None:
    aliases = load_aliases()["indicator_labels"]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Respostas"

    for row, (_canonical, options) in enumerate(aliases.items(), start=5):
        worksheet.cell(row, 3, options[-1])

    found = find_indicator_rows(worksheet, aliases)

    assert set(found) == set(aliases)
    assert found["total_travel_value"] == (5, 3)
    assert all(column == 3 for _, column in found.values())


def test_find_indicator_rows_lists_missing_indicator() -> None:
    aliases = load_aliases()["indicator_labels"]
    workbook = Workbook()
    worksheet = workbook.active

    for row, (canonical, options) in enumerate(aliases.items(), start=1):
        if canonical != "average_ticket":
            worksheet.cell(row, 1, options[0])

    with pytest.raises(DetectionError, match="average_ticket"):
        find_indicator_rows(worksheet, aliases)


def test_header_detector_honors_scan_limit() -> None:
    aliases = load_aliases()["base_columns"]
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(31, 1, "ID Solicitação")
    worksheet.cell(31, 2, "Data Solicitação")

    limited = HeaderDetector().best_header(
        worksheet,
        aliases,
        required_fields=("request_id", "request_date"),
        scan_rows=30,
    )
    expanded = HeaderDetector().best_header(
        worksheet,
        aliases,
        required_fields=("request_id", "request_date"),
        scan_rows=31,
    )

    assert limited[2] == 0
    assert expanded == (31, {"request_id": 1, "request_date": 2}, 2)


def test_workbook_detector_rejects_same_sheet_for_multiple_roles() -> None:
    aliases = load_aliases()
    workbook = Workbook()
    worksheet = workbook.active
    headers = [aliases["base_columns"][field][0] for field in BASE_REQUIRED_FIELDS]
    headers.extend(
        [
            aliases["policy_columns"]["limit_value"][0],
            aliases["policy_columns"]["min_lead_days"][0],
            aliases["response_columns"]["indicator"][0],
            aliases["response_columns"]["answer"][0],
        ]
    )
    worksheet.append(headers)

    with pytest.raises(DetectionError, match="mesma aba"):
        WorkbookDetector(aliases).detect(workbook)


def test_workbook_detector_explains_incomplete_secondary_role() -> None:
    aliases = load_aliases()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Base Completa"
    worksheet.append(
        [aliases["base_columns"][field][0] for field in BASE_REQUIRED_FIELDS]
    )

    with pytest.raises(DetectionError) as error:
        WorkbookDetector(aliases).detect(workbook)

    message = str(error.value)
    assert "policies" in message
    assert "Campos não encontrados" in message

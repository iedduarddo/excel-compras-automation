"""Testes da validação estrutural e numérica do arquivo final."""

from __future__ import annotations

from datetime import date
from pathlib import Path
from unittest.mock import Mock

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill

import src.excel.validation as validation_module
from src.core.exceptions import ValidationError
from src.core.models import SheetLayout, TravelResult, WorkbookLayout
from src.excel.validation import validate_output
from src.settings import load_aliases

BASE_COLUMNS = {
    "request_id": 1,
    "total_value": 2,
    "lead_days": 3,
    "policy_limit": 4,
    "policy_status": 5,
    "limit_difference": 6,
    "priority": 7,
    "priority_score": 8,
}


def make_travel(total_value: float = 500.0) -> TravelResult:
    return TravelResult(
        worksheet_row=2,
        request_id="VIA-001",
        request_date=date(2026, 1, 1),
        travel_date=date(2026, 1, 10),
        service_type="Aéreo",
        supplier="Fornecedor",
        cost_center="CC1001",
        booking_status="Confirmada",
        criticality="Normal",
        card_status="Conferido",
        total_value=total_value,
        lead_days=9,
        policy_limit=1000.0,
        min_lead_days=5,
        policy_status="OK",
        limit_difference=total_value - 1000,
    )


def make_output_file(
    tmp_path: Path,
    *,
    chart: bool = True,
    conditional_rules: int = 2,
    support_state: str = "hidden",
) -> tuple[Path, WorkbookLayout, dict[str, object], list[TravelResult]]:
    tmp_path.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    base = workbook.active
    base.title = "Base"
    for field, column in BASE_COLUMNS.items():
        base.cell(1, column, field)
    base.cell(2, BASE_COLUMNS["request_id"], "VIA-001")
    for field in BASE_COLUMNS:
        if field != "request_id":
            base.cell(2, BASE_COLUMNS[field], "=1")

    for index in range(conditional_rules):
        rule = FormulaRule(
            formula=[f"$G2={index}"],
            fill=PatternFill("solid", fgColor="FFFF0000"),
        )
        base.conditional_formatting.add("A2:H2", rule)

    responses = workbook.create_sheet("Responses")
    aliases = load_aliases()
    for row, options in enumerate(aliases["indicator_labels"].values(), start=1):
        responses.cell(row, 1, options[0])
        responses.cell(row, 2, "=1")

    if chart:
        bar_chart = BarChart()
        bar_chart.add_data(
            Reference(base, min_col=2, min_row=1, max_row=2),
            titles_from_data=True,
        )
        responses.add_chart(bar_chart, "D2")

    workbook.create_sheet("Policies")
    support = workbook.create_sheet("Apoio_Automacao")
    support.sheet_state = support_state

    output_file = tmp_path / "resultado.xlsx"
    workbook.save(output_file)
    workbook.close()

    layout = WorkbookLayout(
        base=SheetLayout("Base", 1, BASE_COLUMNS),
        policies=SheetLayout("Policies", 1, {}),
        responses=SheetLayout(
            "Responses",
            1,
            {"indicator": 1, "answer": 2},
        ),
    )
    return output_file, layout, aliases, [make_travel()]


def run_validation(
    output_file: Path,
    layout: WorkbookLayout,
    aliases: dict[str, object],
    travels: list[TravelResult],
    *,
    native: bool = False,
    cached: bool = False,
) -> dict[str, object]:
    return validate_output(
        output_file=output_file,
        layout=layout,
        aliases=aliases,
        travels=travels,
        last_row=2,
        native_pivot_expected=native,
        cached_values_expected=cached,
    )


def test_validate_output_accepts_complete_fallback_workbook(tmp_path: Path) -> None:
    output_file, layout, aliases, travels = make_output_file(tmp_path)

    checks = run_validation(output_file, layout, aliases, travels)

    assert checks == {
        "travel_rows": 1,
        "formula_cells_checked": 7,
        "indicator_formulas_checked": 10,
        "total_value": None,
        "formula_errors": 0,
        "conditional_formatting_rules": 2,
        "pivot_parts": 0,
        "chart_parts": 1,
        "support_sheet_hidden": True,
    }


@pytest.mark.parametrize(
    ("field", "coordinate"),
    [
        ("total_value", "B2"),
        ("lead_days", "C2"),
        ("priority_score", "H2"),
    ],
)
def test_validate_output_rejects_missing_derived_formula(
    field: str,
    coordinate: str,
    tmp_path: Path,
) -> None:
    output_file, layout, aliases, travels = make_output_file(tmp_path)
    workbook = load_workbook(output_file)
    workbook["Base"].cell(2, BASE_COLUMNS[field], 10)
    workbook.save(output_file)
    workbook.close()

    with pytest.raises(ValidationError, match=coordinate):
        run_validation(output_file, layout, aliases, travels)


def test_validate_output_rejects_indicator_without_formula(tmp_path: Path) -> None:
    output_file, layout, aliases, travels = make_output_file(tmp_path)
    workbook = load_workbook(output_file)
    workbook["Responses"]["B1"] = 500
    workbook.save(output_file)
    workbook.close()

    with pytest.raises(ValidationError, match="total_travel_value"):
        run_validation(output_file, layout, aliases, travels)


def test_validate_output_rejects_cached_formula_error(tmp_path: Path) -> None:
    output_file, layout, aliases, travels = make_output_file(tmp_path)
    workbook = load_workbook(output_file)
    workbook["Apoio_Automacao"]["A1"] = "#REF!"
    workbook.save(output_file)
    workbook.close()

    with pytest.raises(ValidationError, match=r"Apoio_Automacao!A1=#REF!"):
        run_validation(output_file, layout, aliases, travels)


def test_validate_output_requires_chart_and_conditional_rules(tmp_path: Path) -> None:
    no_chart, layout, aliases, travels = make_output_file(
        tmp_path / "sem_grafico",
        chart=False,
    )
    with pytest.raises(ValidationError, match="Nenhum gráfico"):
        run_validation(no_chart, layout, aliases, travels)

    one_rule, layout, aliases, travels = make_output_file(
        tmp_path / "sem_regras",
        conditional_rules=1,
    )
    with pytest.raises(ValidationError, match="formatação condicional"):
        run_validation(one_rule, layout, aliases, travels)


def test_validate_output_requires_native_pivot_only_when_requested(
    tmp_path: Path,
) -> None:
    output_file, layout, aliases, travels = make_output_file(tmp_path)

    run_validation(output_file, layout, aliases, travels, native=False)
    with pytest.raises(ValidationError, match="Tabela Dinâmica nativa"):
        run_validation(output_file, layout, aliases, travels, native=True)


@pytest.mark.parametrize("support_state", ["visible", "veryHidden"])
def test_validate_output_requires_hidden_support_sheet(
    support_state: str,
    tmp_path: Path,
) -> None:
    output_file, layout, aliases, travels = make_output_file(
        tmp_path,
        support_state=support_state,
    )

    with pytest.raises(ValidationError, match="deve permanecer oculta"):
        run_validation(output_file, layout, aliases, travels)


def test_validate_output_requires_support_sheet(tmp_path: Path) -> None:
    output_file, layout, aliases, travels = make_output_file(tmp_path)
    workbook = load_workbook(output_file)
    del workbook["Apoio_Automacao"]
    workbook.save(output_file)
    workbook.close()

    with pytest.raises(ValidationError, match="auxiliar"):
        run_validation(output_file, layout, aliases, travels)


def install_cached_workbooks(
    monkeypatch: pytest.MonkeyPatch,
    output_file: Path,
    aliases: dict[str, object],
    *,
    base_total: object,
    indicator_total: object,
) -> tuple[Workbook, Workbook]:
    formula_workbook = load_workbook(output_file, data_only=False)
    value_workbook = load_workbook(output_file, data_only=False)
    value_workbook["Base"].cell(2, BASE_COLUMNS["total_value"], base_total)
    total_label = aliases["indicator_labels"]["total_travel_value"][0]
    total_row = next(
        cell.row
        for cell in value_workbook["Responses"]["A"]
        if cell.value == total_label
    )
    value_workbook["Responses"].cell(total_row, 2, indicator_total)

    formula_close = Mock()
    value_close = Mock()
    formula_workbook.close = formula_close
    value_workbook.close = value_close

    def fake_load_workbook(
        _: Path,
        *,
        data_only: bool,
        **__: object,
    ) -> Workbook:
        return value_workbook if data_only else formula_workbook

    monkeypatch.setattr(validation_module, "load_workbook", fake_load_workbook)
    return formula_workbook, value_workbook


def test_validate_output_reconciles_cached_totals_and_closes_workbooks(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    output_file, layout, aliases, travels = make_output_file(tmp_path)
    formula_workbook, value_workbook = install_cached_workbooks(
        monkeypatch,
        output_file,
        aliases,
        base_total=500,
        indicator_total=500,
    )

    checks = run_validation(
        output_file,
        layout,
        aliases,
        travels,
        cached=True,
    )

    assert checks["total_value"] == 500.0
    formula_workbook.close.assert_called_once_with()
    value_workbook.close.assert_called_once_with()


def test_validate_output_rejects_mismatched_and_invalid_cached_values(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    output_file, layout, aliases, travels = make_output_file(tmp_path)
    install_cached_workbooks(
        monkeypatch,
        output_file,
        aliases,
        base_total=499,
        indicator_total=500,
    )
    with pytest.raises(ValidationError, match="Excel=499.00"):
        run_validation(output_file, layout, aliases, travels, cached=True)

    formula_workbook, value_workbook = install_cached_workbooks(
        monkeypatch,
        output_file,
        aliases,
        base_total="inválido",
        indicator_total=500,
    )
    with pytest.raises(ValidationError, match=r"Valor Total inválido.*Base!B2"):
        run_validation(output_file, layout, aliases, travels, cached=True)
    formula_workbook.close.assert_called_once_with()
    value_workbook.close.assert_called_once_with()


@pytest.mark.parametrize("indicator_total", [None, "texto", 499])
def test_validate_output_rejects_invalid_total_indicator(
    indicator_total: object,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    output_file, layout, aliases, travels = make_output_file(tmp_path)
    install_cached_workbooks(
        monkeypatch,
        output_file,
        aliases,
        base_total=500,
        indicator_total=indicator_total,
    )

    expected = (
        "Indicador de valor total inválido"
        if indicator_total in {None, "texto"}
        else "não confere"
    )
    with pytest.raises(ValidationError, match=expected):
        run_validation(output_file, layout, aliases, travels, cached=True)

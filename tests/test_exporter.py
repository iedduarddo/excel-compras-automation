"""Testes da exportação de dados para integração ERP."""

from __future__ import annotations

import csv
import hashlib
import json
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from src.core.exceptions import AutomationError
from src.services.exporter import ERPExporter


def create_erp_workbook(path: Path) -> None:
    workbook = Workbook()
    base = workbook.active
    base.title = "Base_Viagens"
    base.append(
        [
            "ID Solicitação",
            "Data Solicitação",
            "Data Viagem",
            "Viajante",
            "Área",
            "Centro de Custo",
            "Tipo de Serviço",
            "Destino",
            "Fornecedor",
            "Quantidade",
            "Valor Unitário",
            "Taxas",
            "Status Reserva",
            "Criticidade",
            "Status Cartão",
            "Valor Total",
            "Status Política",
            "Prioridade",
        ]
    )
    base.append(
        [
            "REQ-001",
            date(2026, 1, 1),
            date(2026, 1, 15),
            "Ana Lima",
            "Financeiro",
            "CC-100",
            "Aéreo",
            "Recife",
            "Fornecedor A",
            2,
            400,
            50,
            "Confirmada",
            "Normal",
            "Conferido",
            "=J2*K2+L2",
            '=IF(P2>1000,"Fora","OK")',
            '="Normal"',
        ]
    )
    base.append(
        [
            "REQ-002",
            date(2026, 1, 10),
            date(2026, 1, 12),
            "Bruno Reis",
            "Comercial",
            "CC-200",
            "Aéreo",
            "Salvador",
            "Fornecedor B",
            1,
            1200,
            None,
            "Pendente",
            "Emergencial",
            "Pendente",
            "=J3*K3+L3",
            '=IF(P3>1000,"Fora","OK")',
            '="Crítica"',
        ]
    )

    policies = workbook.create_sheet("Políticas_Fornecedores")
    policies.append(
        [
            "Tipo de Serviço",
            "Limite por Viajante (R$)",
            "Antecedência Mínima (dias)",
        ]
    )
    policies.append(["Aéreo", 1000, 10])

    responses = workbook.create_sheet("Respostas")
    responses.append(["Indicador", "Resposta"])

    workbook.save(path)
    workbook.close()


def test_constructor_does_not_create_output_directory(tmp_path: Path) -> None:
    output_dir = tmp_path / "erp"

    ERPExporter(tmp_path / "entrada.xlsx", output_dir)

    assert not output_dir.exists()


def test_process_and_export_creates_normalized_artifacts(tmp_path: Path) -> None:
    workbook_path = tmp_path / "viagens.xlsx"
    output_dir = tmp_path / "erp"
    create_erp_workbook(workbook_path)

    result = ERPExporter(workbook_path, output_dir).process_and_export()

    assert result.record_count == 2
    assert result.json_file.name == "erp_carga_compras.json"
    assert result.csv_file.name == "erp_carga_compras.csv"
    assert result.checksum_file.name == "erp_carga_compras.json.sha256"

    records = json.loads(result.json_file.read_text(encoding="utf-8"))
    assert records[0] == {
        "id_solicitacao": "REQ-001",
        "data_solicitacao": "2026-01-01",
        "data_viagem": "2026-01-15",
        "viajante": "Ana Lima",
        "area": "Financeiro",
        "c1_cc": "CC-100",
        "tipo_servico": "Aéreo",
        "destino": "Recife",
        "fornecedor": "Fornecedor A",
        "c1_quant": 2.0,
        "c7_preco": 400.0,
        "taxas": 50.0,
        "c7_total": 850.0,
        "status_reserva": "Confirmada",
        "criticidade": "Normal",
        "status_politica": "OK",
        "prioridade": "Normal",
    }
    assert records[1]["c7_total"] == 1200.0
    assert records[1]["status_politica"] == "Fora"
    assert records[1]["prioridade"] == "Crítica"
    assert all(not str(record["c7_total"]).startswith("=") for record in records)

    with result.csv_file.open(encoding="utf-8-sig", newline="") as file:
        csv_records = list(csv.DictReader(file, delimiter=";"))
    assert [record["id_solicitacao"] for record in csv_records] == [
        "REQ-001",
        "REQ-002",
    ]
    assert csv_records[0]["data_solicitacao"] == "2026-01-01"

    expected_hash = hashlib.sha256(result.json_file.read_bytes()).hexdigest()
    assert result.sha256 == expected_hash
    assert result.checksum_file.read_text(encoding="utf-8") == (
        f"{expected_hash}  {result.json_file.name}\n"
    )


def test_exporter_rejects_missing_or_invalid_workbook(tmp_path: Path) -> None:
    output_dir = tmp_path / "erp"

    with pytest.raises(AutomationError, match="não encontrada"):
        ERPExporter(tmp_path / "ausente.xlsx", output_dir).process_and_export()

    invalid_workbook = tmp_path / "invalida.xlsx"
    invalid_workbook.write_text("não é uma planilha", encoding="utf-8")
    with pytest.raises(AutomationError, match="Não foi possível abrir"):
        ERPExporter(invalid_workbook, output_dir).process_and_export()

    assert not output_dir.exists()


def test_exporter_does_not_overwrite_existing_artifacts(tmp_path: Path) -> None:
    workbook_path = tmp_path / "viagens.xlsx"
    output_dir = tmp_path / "erp"
    create_erp_workbook(workbook_path)
    exporter = ERPExporter(workbook_path, output_dir)
    first_result = exporter.process_and_export()
    original_json = first_result.json_file.read_bytes()

    with pytest.raises(AutomationError, match="não sobrescreve"):
        exporter.process_and_export()

    assert first_result.json_file.read_bytes() == original_json

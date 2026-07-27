"""Testes unitários do motor de políticas e cálculos de viagens."""

from datetime import date, datetime

import pytest
from openpyxl import Workbook
from openpyxl.utils.datetime import to_excel

from src.business.policies import (
    _as_date,
    _as_number,
    analyze_travels,
    find_last_data_row,
    read_policies,
)
from src.core.exceptions import ValidationError
from src.core.models import Policy, SheetLayout, WorkbookLayout
from src.services.text import normalize_text

BASE_COLUMNS = {
    "request_id": 1,
    "request_date": 2,
    "travel_date": 3,
    "service_type": 4,
    "supplier": 5,
    "cost_center": 6,
    "booking_status": 7,
    "criticality": 8,
    "card_status": 9,
    "quantity": 10,
    "unit_value": 11,
    "fees": 12,
}

POLICY_COLUMNS = {
    "service_type": 1,
    "limit_value": 2,
    "min_lead_days": 3,
}


def make_workbook() -> tuple[Workbook, WorkbookLayout]:
    """Cria uma estrutura mínima totalmente em memória."""

    workbook = Workbook()
    base = workbook.active
    base.title = "Base"
    base.append(list(BASE_COLUMNS))

    policies = workbook.create_sheet("Policies")
    policies.append(list(POLICY_COLUMNS))

    workbook.create_sheet("Responses")
    layout = WorkbookLayout(
        base=SheetLayout("Base", 1, BASE_COLUMNS),
        policies=SheetLayout("Policies", 1, POLICY_COLUMNS),
        responses=SheetLayout("Responses", 1, {}),
    )
    return workbook, layout


def append_travel(
    workbook: Workbook,
    *,
    request_id: object = "REQ-001",
    request_date: object = date(2026, 1, 1),
    travel_date: object = date(2026, 1, 11),
    service_type: object = "Aéreo Nacional",
    supplier: object = "Fornecedor",
    cost_center: object = "CC1001",
    booking_status: object = "Confirmada",
    criticality: object = "Normal",
    card_status: object = "Conferido",
    quantity: object = 1,
    unit_value: object = 500,
    fees: object = 0,
) -> None:
    """Adiciona uma solicitação usando a ordem definida em ``BASE_COLUMNS``."""

    workbook["Base"].append(
        [
            request_id,
            request_date,
            travel_date,
            service_type,
            supplier,
            cost_center,
            booking_status,
            criticality,
            card_status,
            quantity,
            unit_value,
            fees,
        ]
    )


def make_policy(
    *,
    service_type: str = "Aéreo Nacional",
    limit_value: float = 500.0,
    min_lead_days: int = 5,
) -> dict[str, Policy]:
    policy = Policy(
        service_type=service_type,
        limit_value=limit_value,
        min_lead_days=min_lead_days,
        worksheet_row=2,
    )
    return {normalize_text(service_type): policy}


def test_read_policies_skips_empty_and_malformed_rows() -> None:
    workbook, layout = make_workbook()
    worksheet = workbook["Policies"]
    worksheet.append([None, None, None])
    worksheet.append(["  Aéreo Nacional  ", 1800, 10])
    worksheet.append(["Linha incompleta", "sem limite", 5])
    worksheet.append(["Hotel", 900.5, 7.9])

    policies = read_policies(workbook, layout)

    assert set(policies) == {"aereo nacional", "hotel"}
    assert policies["aereo nacional"] == Policy(
        service_type="Aéreo Nacional",
        limit_value=1800.0,
        min_lead_days=10,
        worksheet_row=3,
    )
    assert policies["hotel"].min_lead_days == 7
    assert policies["hotel"].worksheet_row == 5


def test_read_policies_rejects_normalized_duplicates() -> None:
    workbook, layout = make_workbook()
    workbook["Policies"].append(["Aéreo-Nacional", 1800, 10])
    workbook["Policies"].append([" aéreo nacional ", 2000, 12])

    with pytest.raises(ValidationError, match="duplicada"):
        read_policies(workbook, layout)


def test_read_policies_requires_at_least_one_valid_policy() -> None:
    workbook, layout = make_workbook()
    workbook["Policies"].append([None, None, None])
    workbook["Policies"].append(["Hotel", "valor inválido", 7])

    with pytest.raises(ValidationError, match="Nenhuma política válida"):
        read_policies(workbook, layout)


def test_find_last_data_row_ignores_trailing_empty_identifiers() -> None:
    workbook, layout = make_workbook()
    append_travel(workbook, request_id="REQ-001")
    append_travel(workbook, request_id=None)
    append_travel(workbook, request_id=" REQ-003 ")
    append_travel(workbook, request_id="   ")

    assert find_last_data_row(workbook, layout) == 4


def test_find_last_data_row_rejects_base_without_requests() -> None:
    workbook, layout = make_workbook()
    append_travel(workbook, request_id=None)
    append_travel(workbook, request_id=" ")

    with pytest.raises(ValidationError, match="não possui solicitações"):
        find_last_data_row(workbook, layout)


def test_analyze_travels_applies_all_policy_outcomes() -> None:
    workbook, layout = make_workbook()
    policies = make_policy()
    append_travel(
        workbook,
        request_id=" REQ-OK ",
        travel_date=date(2026, 1, 6),
        supplier=" Fornecedor A ",
        cost_center=" CC1001 ",
        booking_status=" Confirmada ",
        criticality=" Normal ",
        card_status=" Conferido ",
        quantity=2,
        unit_value=200,
        fees=100,
    )
    append_travel(
        workbook,
        request_id="REQ-CUSTO",
        travel_date=date(2026, 1, 20),
        unit_value=500.01,
        fees=None,
    )
    append_travel(
        workbook,
        request_id="REQ-PRAZO",
        travel_date=date(2026, 1, 5),
        unit_value=100,
        fees="",
    )
    append_travel(
        workbook,
        request_id=None,
        service_type="Aéreo Nacional",
    )
    append_travel(
        workbook,
        request_id="REQ-REVISAR",
        service_type="Serviço sem política",
        supplier=None,
        cost_center=None,
        booking_status=None,
        criticality=None,
        card_status=None,
        quantity=3,
        unit_value=40.185,
        fees=5,
    )

    results = analyze_travels(workbook, layout, policies)
    by_id = {result.request_id: result for result in results}

    assert set(by_id) == {"REQ-OK", "REQ-CUSTO", "REQ-PRAZO", "REQ-REVISAR"}

    compliant = by_id["REQ-OK"]
    assert compliant.total_value == 500.0
    assert compliant.lead_days == 5
    assert compliant.policy_limit == 500.0
    assert compliant.min_lead_days == 5
    assert compliant.policy_status == "OK"
    assert compliant.limit_difference == 0.0
    assert compliant.supplier == "Fornecedor A"
    assert compliant.cost_center == "CC1001"
    assert compliant.booking_status == "Confirmada"
    assert compliant.criticality == "Normal"
    assert compliant.card_status == "Conferido"

    assert by_id["REQ-CUSTO"].policy_status == "Fora"
    assert by_id["REQ-CUSTO"].limit_difference == 0.01
    assert by_id["REQ-PRAZO"].policy_status == "Fora"
    assert by_id["REQ-PRAZO"].lead_days == 4

    without_policy = by_id["REQ-REVISAR"]
    assert without_policy.total_value == 125.56
    assert without_policy.policy_status == "Revisar"
    assert without_policy.policy_limit == 0.0
    assert without_policy.min_lead_days == 0
    assert without_policy.limit_difference == 125.56
    assert without_policy.supplier == ""
    assert without_policy.cost_center == ""


def test_analyze_travels_accepts_excel_serial_dates() -> None:
    workbook, layout = make_workbook()
    append_travel(
        workbook,
        request_date=to_excel(datetime(2026, 2, 1)),
        travel_date=to_excel(datetime(2026, 2, 11)),
    )

    [result] = analyze_travels(workbook, layout, make_policy())

    assert result.request_date == datetime(2026, 2, 1)
    assert result.travel_date == datetime(2026, 2, 11)
    assert result.lead_days == 10


@pytest.mark.parametrize(
    ("field", "value", "expected_name"),
    [
        ("quantity", True, "Quantidade"),
        ("unit_value", "500", "Valor Unitário"),
        ("fees", False, "Taxas"),
    ],
)
def test_analyze_travels_rejects_invalid_numeric_values(
    field: str,
    value: object,
    expected_name: str,
) -> None:
    workbook, layout = make_workbook()
    values = {
        "quantity": 1,
        "unit_value": 500,
        "fees": 0,
    }
    values[field] = value
    append_travel(workbook, **values)

    with pytest.raises(ValidationError, match=expected_name):
        analyze_travels(workbook, layout, make_policy())


@pytest.mark.parametrize(
    ("field", "value", "expected_name"),
    [
        ("request_date", "ontem", "Data Solicitação"),
        ("travel_date", "amanhã", "Data Viagem"),
    ],
)
def test_analyze_travels_rejects_invalid_dates(
    field: str,
    value: object,
    expected_name: str,
) -> None:
    workbook, layout = make_workbook()
    values = {
        "request_date": date(2026, 1, 1),
        "travel_date": date(2026, 1, 11),
    }
    values[field] = value
    append_travel(workbook, **values)

    with pytest.raises(ValidationError, match=expected_name):
        analyze_travels(workbook, layout, make_policy())


def test_conversion_helpers_preserve_valid_values_and_explain_errors() -> None:
    input_date = date(2026, 3, 10)
    input_datetime = datetime(2026, 3, 10, 8, 30)

    assert _as_number(12, "Quantidade", 7) == 12.0
    assert _as_date(input_date, "Data", 7) is input_date
    assert _as_date(input_datetime, "Data", 7) is input_datetime

    with pytest.raises(ValidationError, match=r"Linha 7.*Quantidade"):
        _as_number([], "Quantidade", 7)

    with pytest.raises(ValidationError, match=r"Linha 8.*Data Viagem"):
        _as_date("2026/03/10", "Data Viagem", 8)
